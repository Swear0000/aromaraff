import asyncio
from aiogram import Bot, Dispatcher, Router, types, F
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, FSInputFile
import aiosqlite
import smtplib
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
import os
import json
from datetime import datetime
from aiohttp import web
from openpyxl.styles import Font, Alignment, Border, Side  # Добавьте эту строку
import logging

# Настройки
BOT_TOKEN = '8034399145:AAEVIsikLZDVD3aGMJ8cDZaeTN91VOivAHM'  # Вставь токен от @BotFather
YOUR_ADMIN_ID = 6329978401  # Вставь свой Telegram ID
ADMIN_EMAIL = 'swear000@yandex.ru'  # Вставь свой Яндекс email
EMAIL_PASSWORD = 'cobrcbopfkzzfisr'  # Вставь пароль приложения от Яндекса

# Константы для очистки корзины
CART_LIFETIME_DAYS = 14  # Корзина хранится 2 недели
CLEANUP_INTERVAL_HOURS = 24  # Проверка и очистка старых корзин раз в 24 часа

# Константы для пагинации
PRODUCTS_PER_PAGE = 5

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()
router = Router()

# Регистрация роутера
dp.include_router(router)

# --- Состояния для FSM ---
class AdminStates(StatesGroup):
    add_product_name = State()
    add_product_price = State()
    add_product_description = State()
    add_product_quantity = State()
    add_product_image = State()
    add_product_category = State()
    edit_product_select = State()
    edit_product_menu = State()
    edit_product_name_input = State()
    edit_product_price_input = State()
    edit_product_description_input = State()
    edit_product_category_input = State()
    edit_product_image_input = State()
    # НОВЫЕ СОСТОЯНИЯ ДЛЯ ИМПОРТА
    delete_product_confirmation = State() # <-- ДОБАВЬТЕ ЭТУ СТРОКУ
    manage_categories_menu = State()
    add_category_name = State()
    rename_category_name = State()
    manage_product_characteristics = State()
    add_characteristic_name = State()
    add_characteristic_value = State()
    add_characteristic_price = State()
    add_characteristic_quantity = State()
    select_characteristic_to_edit = State()
    edit_characteristic_field = State()
    edit_characteristic_price_input = State()
    edit_characteristic_quantity_input = State()
    

class UserStates(StatesGroup):
    awaiting_delivery_name = State()
    awaiting_delivery_address = State()
    awaiting_delivery_phone = State()
    awaiting_order_confirmation = State()
    cart_remove_select_item = State()
    cart_change_qty_enter_new_qty = State()
    viewing_category_page = State()
    awaiting_user_message_to_admin = State()


# --- Вспомогательные функции для работы с корзиной в БД ---
async def load_user_cart_from_db(user_id):
    """Загружает корзину пользователя из базы данных."""
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT cart_items FROM user_carts WHERE user_id = ?", (user_id,))
        result = await cursor.fetchone()
        if result and result[0]:
            print(f"DEBUG: Loaded cart for user {user_id}: {result[0]}")
            return json.loads(result[0])
        print(f"DEBUG: No cart found or empty cart for user {user_id}")
        return []

async def save_user_cart_to_db(user_id, cart):
    """Сохраняет корзину пользователя в базу данных и обновляет метку времени."""
    try:
        async with aiosqlite.connect("products.db") as db:
            cart_json = json.dumps(cart)
            current_time = datetime.now().strftime('%d-%m-%Y')
            await db.execute(
                "INSERT OR REPLACE INTO user_carts (user_id, cart_items, last_updated) VALUES (?, ?, ?)",
                (user_id, cart_json, current_time)
            )
            await db.commit()
            print(f"DEBUG: Saved cart for user {user_id}: {cart_json} at {current_time}")
    except Exception as e:
        print(f"ERROR in save_user_cart_to_db: {str(e)}")
        raise


async def init_db():
    async with aiosqlite.connect("products.db") as db:
        
        # Таблица категорий
        await db.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE
            )""")
        
        
        # Обновленная таблица products (без price и quantity)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                description TEXT,
                category_id INTEGER,
                image_url TEXT,
                FOREIGN KEY (category_id) REFERENCES categories (id)
            )""")

        # НОВАЯ ТАБЛИЦА: product_characteristics
        await db.execute("""
            CREATE TABLE IF NOT EXISTS product_characteristics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                name TEXT NOT NULL,         -- Например, "Объем", "Цвет"
                value TEXT NOT NULL,        -- Например, "5мл", "Красный"
                price INTEGER NOT NULL,     -- Цена для этой характеристики
                quantity INTEGER NOT NULL,  -- Количество для этой характеристики
                FOREIGN KEY (product_id) REFERENCES products (id) ON DELETE CASCADE
            )""")
        
        # Создание таблицы orders с полем created_at
        await db.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                delivery_info TEXT,
                status TEXT DEFAULT 'Оформлен',
                created_at TEXT
            )""")
        
        
        await db.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                characteristic_id INTEGER,  
                quantity INTEGER,
                FOREIGN KEY (order_id) REFERENCES orders (id),
                FOREIGN KEY (characteristic_id) REFERENCES product_characteristics (id) ON DELETE CASCADE
            )""")
        
        # Таблица для избранных товаров (пока оставим на уровне product_id для простоты)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS user_favorites (
                user_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, product_id),
                FOREIGN KEY (product_id) REFERENCES products (id) ON DELETE CASCADE
            )""")

        # МИГРАЦИЯ СХЕМЫ: Удаление старых колонок 'price' и 'quantity' из 'products'
        cursor = await db.execute("PRAGMA table_info(products)")
        columns = [col[1] for col in await cursor.fetchall()]
        
        if 'price' in columns:
            print("DEBUG: 'price' column found in 'products'. Attempting to drop it.")
            # SQLite не поддерживает прямое DROP COLUMN, нужна миграция
            await db.execute("ALTER TABLE products RENAME TO _old_products_with_price")
            await db.execute("""
                CREATE TABLE products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT,
                    description TEXT,
                    category_id INTEGER,
                    image_url TEXT
                )""")
            await db.execute("""
                INSERT INTO products (id, name, description, category_id, image_url)
                SELECT id, name, description, category_id, image_url FROM _old_products_with_price
            """)
            await db.execute("DROP TABLE _old_products_with_price")
            print("DEBUG: 'price' column removed and table recreated.")

        if 'quantity' in columns:
            print("DEBUG: 'quantity' column found in 'products'. Attempting to drop it.")
            # Повторная миграция, если 'price' уже удаляли, или первая, если только 'quantity'
            cursor_new = await db.execute("PRAGMA table_info(products)")
            current_columns = [col[1] for col in await cursor_new.fetchall()]
            if 'quantity' in current_columns: # Проверяем еще раз, вдруг уже удалилась с price
                await db.execute("ALTER TABLE products RENAME TO _old_products_with_quantity")
                await db.execute("""
                    CREATE TABLE products (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT,
                        description TEXT,
                        category_id INTEGER,
                        image_url TEXT
                    )""")
                await db.execute("""
                    INSERT INTO products (id, name, description, category_id, image_url)
                    SELECT id, name, description, category_id, image_url FROM _old_products_with_quantity
                """)
                await db.execute("DROP TABLE _old_products_with_quantity")
                print("DEBUG: 'quantity' column removed and table recreated.")

        # Проверка и добавление колонки category_id в products, если её нет (старая проверка, можно оставить)
        cursor_final = await db.execute("PRAGMA table_info(products)")
        final_columns = [col[1] for col in await cursor_final.fetchall()]
        if 'category_id' not in final_columns:
            print("DEBUG: 'category_id' column missing in 'products'. Adding it.")
            await db.execute("ALTER TABLE products ADD COLUMN category_id INTEGER")
            print("DEBUG: 'category_id' column added to 'products'.")
        
        # Проверка и добавление колонки image_url в products, если её нет (старая проверка, можно оставить)
        if 'image_url' not in final_columns:
            print("DEBUG: 'image_url' column missing in 'products'. Adding it.")
            await db.execute("ALTER TABLE products ADD COLUMN image_url TEXT")
            print("DEBUG: 'image_url' column added to 'products'.")
        
        # Проверка наличия таблицы user_carts и колонки last_updated (оставить как есть)
        cursor = await db.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_carts'")
        user_carts_exists = await cursor.fetchone()
        if not user_carts_exists:
            print("DEBUG: Creating new 'user_carts' table with 'last_updated' column.")
            await db.execute("""
                CREATE TABLE user_carts (
                    user_id INTEGER PRIMARY KEY,
                    cart_items TEXT,
                    last_updated TEXT DEFAULT CURRENT_TIMESTAMP
                )""")
        else:
            cursor = await db.execute("PRAGMA table_info(user_carts)")
            columns = [col[1] for col in await cursor.fetchall()]
            if 'last_updated' not in columns:
                print("DEBUG: 'last_updated' column missing in 'user_carts'. Performing schema migration.")
                await db.execute("ALTER TABLE user_carts RENAME TO _old_user_carts")
                await db.execute("""
                    CREATE TABLE user_carts (
                        user_id INTEGER PRIMARY KEY,
                        cart_items TEXT,
                        last_updated TEXT DEFAULT CURRENT_TIMESTAMP
                    )""")
                await db.execute("""
                    INSERT INTO user_carts (user_id, cart_items, last_updated)
                    SELECT user_id, cart_items, CURRENT_TIMESTAMP FROM _old_user_carts
                """)
                await db.execute("DROP TABLE _old_user_carts")
                print("DEBUG: Schema migration for 'user_carts' complete.")

        # Добавляем стандартные категории, если их нет (оставить как есть)
        initial_categories = ["Женский парфюм", "Мужской парфюм", "Унисекс парфюм"]
        for category_name in initial_categories:
            await db.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (category_name,))
        
       # Проверяем, есть ли колонка created_at
        cursor = await db.execute("PRAGMA table_info(orders)")
        columns = [col[1] for col in await cursor.fetchall()]
        if 'created_at' not in columns:
            print("DEBUG: Migrating orders table to add created_at column")
            # Создаем новую таблицу с нужной структурой
            await db.execute("""
                CREATE TABLE orders_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    delivery_info TEXT,
                    status TEXT DEFAULT 'Оформлен',
                    created_at TEXT
                )""")
        # Копируем данные из старой таблицы в новую, устанавливая текущую дату для created_at
            await db.execute("""
                INSERT INTO orders_new (id, user_id, delivery_info, status, created_at)
                SELECT id, user_id, delivery_info, status, datetime('now')
                FROM orders
            """)
          # Удаляем старую таблицу
            await db.execute("DROP TABLE orders")
            
            # Переименовываем новую таблицу
            await db.execute("ALTER TABLE orders_new RENAME TO orders")
            
            print("DEBUG: Migration completed successfully")

        await db.commit()


async def get_category_name(category_id):
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name FROM categories WHERE id = ?", (category_id,))
        result = await cursor.fetchone()
        return result[0] if result else "Без категории"

# <--- НОВЫЙ БЛОК: Функции для работы с избранным
async def is_product_in_favorites(user_id: int, product_id: int) -> bool:
    """Проверяет, находится ли товар в избранном у пользователя."""
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute(
            "SELECT 1 FROM user_favorites WHERE user_id = ? AND product_id = ?",
            (user_id, product_id)
        )
        return await cursor.fetchone() is not None

async def toggle_favorite_status(user_id: int, product_id: int) -> bool:
    """Добавляет или удаляет товар из избранного. Возвращает True, если товар теперь в избранном."""
    is_currently_favorite = await is_product_in_favorites(user_id, product_id)
    async with aiosqlite.connect("products.db") as db:
        if is_currently_favorite:
            await db.execute(
                "DELETE FROM user_favorites WHERE user_id = ? AND product_id = ?",
                (user_id, product_id)
            )
            await db.commit()
            return False
        else:
            cursor = await db.execute("SELECT 1 FROM products WHERE id = ?", (product_id,))
            if await cursor.fetchone():
                await db.execute(
                    "INSERT OR IGNORE INTO user_favorites (user_id, product_id) VALUES (?, ?)",
                    (user_id, product_id)
                )
                await db.commit()
                return True
            return False

async def send_email(subject, body):
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = ADMIN_EMAIL
    msg['To'] = ADMIN_EMAIL
    try:
        with smtplib.SMTP_SSL('smtp.yandex.ru', 465) as server:
            server.login(ADMIN_EMAIL, EMAIL_PASSWORD)
            server.send_message(msg)
    except Exception as e:
        print(f"Email error: {e}")

async def notify_admin(message_text):
    try:
        await bot.send_message(YOUR_ADMIN_ID, message_text)
    except Exception as e:
        print(f"Telegram notify error: {e}")

# --- Фоновая задача для очистки старых корзин ---
async def clear_old_carts_task():
    while True:
        await asyncio.sleep(CLEANUP_INTERVAL_HOURS * 3600)
        print("DEBUG: Running old carts cleanup task...")
        async with aiosqlite.connect("products.db") as db:
            threshold_time = datetime.now() - timedelta(days=CART_LIFETIME_DAYS)
            threshold_time_str = threshold_time.strftime('%d-%m-%Y')
            await db.execute("DELETE FROM user_carts WHERE last_updated < ?", (threshold_time_str,))
            await db.commit()
            print(f"DEBUG: Old carts cleanup complete. Removed carts older than {threshold_time_str}")

# --- Основные команды и кнопки ---
@router.message(Command("start"))
async def start(message: types.Message):
    # Базовые кнопки для всех пользователей
    keyboard = [
        [KeyboardButton(text="📋 Каталог"), KeyboardButton(text="🛒 Корзина")],
        [KeyboardButton(text="👤 Личный кабинет"), KeyboardButton(text="❓ Помощь")]
    ]
    
    # Добавляем кнопку админ-панели, если пользователь является админом
    if message.from_user and message.from_user.id == YOUR_ADMIN_ID:
        keyboard.append([KeyboardButton(text="⚙️ ЛК - Админа")])
    
    markup = ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    await message.answer("Добро пожаловать!", reply_markup=markup)

@router.message(F.text == "👤 Личный кабинет")
async def personal_account(message: types.Message):
    keyboard = [
        [KeyboardButton(text="📦 Мои заказы")],
        [KeyboardButton(text="❤️ Избранное")],  # Добавляем кнопку избранного
        [KeyboardButton(text="↩️ Главное меню")]
    ]
    await message.answer("Вы в личном кабинете.", reply_markup=ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True))
    
@router.message(F.text == "❤️ Избранное")
async def show_favorites(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    
    async with aiosqlite.connect("products.db") as db:
        # Получаем избранные товары пользователя с их характеристиками
        cursor = await db.execute("""
            SELECT DISTINCT
                p.id,
                p.name,
                p.description,
                p.image_url,
                (
                    SELECT COUNT(*)
                    FROM product_characteristics pc
                    WHERE pc.product_id = p.id
                ) as has_characteristics
            FROM products p
            JOIN user_favorites uf ON p.id = uf.product_id
            WHERE uf.user_id = ?
        """, (user_id,))
        
        favorites = await cursor.fetchall()
        
        if not favorites:
            markup = InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="📋 Перейти в каталог", callback_data="show_main_categories")
            ]])
            await message.answer(
                "У вас пока нет товаров в избранном. "
                "Добавляйте товары в избранное, нажимая на кнопку ❤️ в карточке товара!",
                reply_markup=markup
            )
            return

        await message.answer("**Ваши избранные товары:**", parse_mode="Markdown")

        # Отправляем каждый товар отдельным сообщением
        for prod_id, name, description, image_url, has_characteristics in favorites:
            caption = f"**{name}**\n\n{description or 'Описание отсутствует'}"
            
            # Проверяем, есть ли у товара характеристики
            has_chars = has_characteristics > 0
            
            # Получаем статус избранного (должен быть True, так как это список избранного)
            is_fav = True
            
            # Создаем разметку кнопок
            markup = get_product_card_markup(prod_id, has_chars, is_fav)

            if not has_chars:
                caption += "\n\n**Нет доступных характеристик**"
            
            try:
                if image_url:
                    await bot.send_photo(
                        chat_id=user_id,
                        photo=image_url,
                        caption=caption,
                        reply_markup=markup,
                        parse_mode="Markdown"
                    )
                else:
                    await bot.send_message(
                        chat_id=user_id,
                        text=caption + "\n\n(Изображение отсутствует)",
                        reply_markup=markup,
                        parse_mode="Markdown"
                    )
            except Exception as e:
                print(f"Error sending favorite product {prod_id}: {e}")
                await bot.send_message(
                    chat_id=user_id,
                    text=caption + "\n\n(Не удалось загрузить изображение)",
                    reply_markup=markup,
                    parse_mode="Markdown"
                )

        # Добавляем кнопку возврата в личный кабинет
        back_markup = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="📦 Мои заказы")],
                [KeyboardButton(text="❤️ Избранное")],
                [KeyboardButton(text="↩️ Главное меню")]
            ],
            resize_keyboard=True
        )
        await message.answer(
            "Для возврата используйте кнопки ниже:",
            reply_markup=back_markup
        )    
    
    

@router.message(F.text == "⚙️ ЛК - Админа", F.from_user.id == YOUR_ADMIN_ID)
async def admin_account(message: types.Message):
    print("DEBUG: Admin panel accessed") # Добавляем отладку
    keyboard = [
        [KeyboardButton(text="📥 Заказы (админ)"), KeyboardButton(text="📤 Экспорт заказов")],
        [KeyboardButton(text="➕ Добавить товар"), KeyboardButton(text="🎛️ Упр. характеристиками")],
        [KeyboardButton(text="✏️ Редактировать товар"), KeyboardButton(text="🗂️ Категории товаров")],
        [KeyboardButton(text="📤 Экспорт товаров"), KeyboardButton(text="📥 Импорт товаров")],
        [KeyboardButton(text="↩️ Главное меню")]
    ]
    markup = ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    print("DEBUG: Creating admin panel keyboard") # Добавляем отладку
    await message.answer("Вы в админ-панели.", reply_markup=markup)


@router.message(F.text == "↩️ Главное меню")
async def back_to_main_menu(message: types.Message):
    await start(message)

@router.message(F.text == "❓ Помощь")
async def help_command(message: types.Message):
    help_text = (
        "Здравствуйте! Если у вас возникли вопросы по работе бота, товарам или заказам можете связаться с нами!\n\n"
        "Телефон/Telegram: +7993 948 0909\n\n"
        "Часы работы: Пн-Вск с 8.00 до 18.00 по Мск\n\n"
        "Мы всегда готовы помочь!"
    )
    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Написать", callback_data="ask_admin")],
        [InlineKeyboardButton(text="Главное меню", callback_data="back_to_main_from_help")]
    ])
    await message.answer(help_text, reply_markup=markup, parse_mode="Markdown")

@router.callback_query(F.data == "ask_admin")
async def ask_admin_callback(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(UserStates.awaiting_user_message_to_admin)
    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_help_menu")]
    ])
    await callback.message.edit_text("Введите ваше сообщение для администратора:", reply_markup=markup)
    await callback.answer()

@router.message(UserStates.awaiting_user_message_to_admin)
async def process_user_message_to_admin(message: types.Message, state: FSMContext):
    user_name = message.from_user.full_name
    user_id = message.from_user.id
    admin_message = (
        f"Новое сообщение от пользователя {user_name} (ID: {user_id}):\n\n"
        f"{message.text}"
    )
    await notify_admin(admin_message)
    await message.answer("Ваше сообщение отправлено администратору. Мы свяжемся с вами в ближайшее время!")
    await state.clear()
    await help_command(message) # Возвращаем в меню "Помощь"

@router.callback_query(F.data == "back_to_help_menu")
async def back_to_help_menu_callback(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete() # Удаляем предыдущее сообщение с полем ввода
    await help_command(callback.message)
    await callback.answer()

@router.callback_query(F.data == "back_to_main_from_help")
async def back_to_main_from_help_callback(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete() # Удаляем сообщение с меню помощи
    await start(callback.message)
    await callback.answer()

# --- Логика каталога (полностью переработана) ---

@router.message(F.text == "📋 Каталог")
async def show_categories(message: types.Message, state: FSMContext):
    await state.clear()
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("""
            SELECT c.id, c.name, (SELECT COUNT(*) FROM products p WHERE p.category_id = c.id) as product_count
            FROM categories c
        """)
        categories_with_counts = await cursor.fetchall()
    
    buttons = []
    if categories_with_counts:
        for cat_id, cat_name, product_count in categories_with_counts:
            buttons.append([InlineKeyboardButton(text=f"{cat_name} ({product_count})", callback_data=f"show_cat_{cat_id}_page_0")])
    
    # <--- ИЗМЕНЕНО: Добавлена кнопка "Избранное"
    buttons.append([InlineKeyboardButton(text="❤️ Избранное", callback_data="show_fav_page_0")])

    await message.answer("Выберите категорию товаров или просмотрите избранное:", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))

# --- Генераторы клавиатур для карточек (Обновленная) ---
def get_product_card_markup(product_id: int, has_characteristics: bool, is_favorite: bool) -> InlineKeyboardMarkup:
    """
    Создает стандартную клавиатуру для карточки товара.
    has_characteristics: True, если у товара есть характеристики.
    """
    favorite_text = "💔 Удалить из избранного" if is_favorite else "❤️ Добавить в избранное"
    favorite_callback = f"fav_rem_{product_id}" if is_favorite else f"fav_add_{product_id}"

    buttons = []
    if has_characteristics:
        buttons.append([InlineKeyboardButton(text="Выбрать параметры", callback_data=f"select_char_{product_id}")])
    else:
        # Если нет характеристик, возможно, товар должен быть "недоступен" или иметь стандартную цену/кол-во
        # Для простоты, здесь мы не добавляем кнопку "Купить", если нет характеристик
        pass
        
    buttons.append([InlineKeyboardButton(text=favorite_text, callback_data=favorite_callback)])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def get_characteristic_selection_markup(product_id: int, characteristics: list) -> InlineKeyboardMarkup:
    """Создает клавиатуру для выбора характеристики товара."""
    buttons = []
    for char_id, char_name, char_value, char_price, char_qty in characteristics:
        if char_qty > 0: # Показываем только те характеристики, которые есть в наличии
            buttons.append([InlineKeyboardButton(text=f"Купить - {char_name}: {char_value} - {char_price / 100:.2f}₽", callback_data=f"buy_char_{char_id}")])
        else:
            buttons.append([InlineKeyboardButton(text=f"{char_name}: {char_value} (Нет в наличии)", callback_data="noop_char_no_stock")])

    buttons.append([InlineKeyboardButton(text="⬅️ Назад к товару", callback_data=f"back_to_product_view_{product_id}")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def get_quantity_selector_markup(characteristic_id: int, current_qty: int, max_qty: int) -> InlineKeyboardMarkup:
    """Создает клавиатуру для выбора количества для выбранной характеристики."""
    buttons = [
        [
            InlineKeyboardButton(text="-", callback_data=f"qty-char-change_dec_{characteristic_id}_{current_qty}"),
            InlineKeyboardButton(text=f"{current_qty} шт.", callback_data="noop"),
            InlineKeyboardButton(text="+", callback_data=f"qty-char-change_inc_{characteristic_id}_{current_qty}")
        ],
        [InlineKeyboardButton(text=f"✅ Добавить в корзину", callback_data=f"cart_add_char_{characteristic_id}_{current_qty}")],
        [InlineKeyboardButton(text="⬅️ Назад к выбору характеристики", callback_data=f"cancel_qty_selector_char_{characteristic_id}")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


async def send_products_page(message: types.Message, page: int, state: FSMContext, category_id: int = None, from_favorites: bool = False):
    offset = page * PRODUCTS_PER_PAGE
    user_id = message.chat.id
    
    async with aiosqlite.connect("products.db") as db:
        if from_favorites:
            cursor_count = await db.execute("SELECT COUNT(*) FROM user_favorites WHERE user_id = ?", (user_id,))
            total_products = (await cursor_count.fetchone())[0]
            cursor = await db.execute(
                """SELECT p.id, p.name, p.description, p.image_url
                   FROM products p JOIN user_favorites uf ON p.id = uf.product_id
                   WHERE uf.user_id = ? LIMIT ? OFFSET ?""",
                (user_id, PRODUCTS_PER_PAGE, offset)
            )
            products = await cursor.fetchall()
            page_title = f"❤️ Избранное (Страница {page + 1})"
            base_callback = "show_fav"
        else:
            cursor_count = await db.execute("SELECT COUNT(*) FROM products WHERE category_id = ?", (category_id,))
            total_products = (await cursor_count.fetchone())[0]
            cursor = await db.execute(
                "SELECT id, name, description, image_url FROM products WHERE category_id = ? LIMIT ? OFFSET ?",
                (category_id, PRODUCTS_PER_PAGE, offset)
            )
            products = await cursor.fetchall()
            category_name = await get_category_name(category_id)
            page_title = f"📋 Каталог - {category_name} (Страница {page + 1})"
            base_callback = f"show_cat_{category_id}"

    if not products:
        empty_message = "Ваше избранное пока пусто 💔. Добавьте сюда товары, которые вам понравились, чтобы не потерять их и легко вернуться к покупкам! Загляните в наш 📋 Каталог — там точно найдётся что-то особенное для вас!" if from_favorites else "В этой категории пока нет товаров."
        await message.answer(empty_message, 
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="📋 Каталог", callback_data="show_main_categories")]]))
        return

    await message.answer(f"**{page_title}**", parse_mode="Markdown")

    for prod_id, name, desc, image_url in products: # Изменено: нет price, qty
        caption = f"**{name}**\n\n{desc or 'Описание отсутствует'}"
        
        # Проверяем, есть ли у товара характеристики
        async with aiosqlite.connect("products.db") as db_inner:
            cursor_char = await db_inner.execute("SELECT COUNT(*) FROM product_characteristics WHERE product_id = ?", (prod_id,))
            has_characteristics = (await cursor_char.fetchone())[0] > 0
        
        is_fav = await is_product_in_favorites(user_id, prod_id)
        markup = get_product_card_markup(prod_id, has_characteristics, is_fav)

        if not has_characteristics: # Если нет характеристик, считаем товар недоступным для покупки
            caption += "\n\n**Нет доступных характеристик**"
            markup = None # Убираем кнопки, если нельзя купить
            
        if image_url:
            try:
                await bot.send_photo(chat_id=user_id, photo=image_url, caption=caption, reply_markup=markup, parse_mode="Markdown")
            except Exception as e:
                await bot.send_message(user_id, caption + "\n\n(Не удалось загрузить изображение)", reply_markup=markup, parse_mode="Markdown")
        else:
            await bot.send_message(user_id, caption + "\n\n(Изображение отсутствует)", reply_markup=markup, parse_mode="Markdown")

    nav_buttons = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton(text="⬅️ Назад", callback_data=f"{base_callback}_page_{page - 1}"))
    if (page + 1) * PRODUCTS_PER_PAGE < total_products:
        nav_buttons.append(InlineKeyboardButton(text="Далее ➡️", callback_data=f"{base_callback}_page_{page + 1}"))

    if nav_buttons:
        pagination_markup = InlineKeyboardMarkup(inline_keyboard=[nav_buttons])
        await message.answer("Перейти на другую страницу:", reply_markup=pagination_markup)
 
@router.callback_query(F.data.startswith("show_cat_"))
async def navigate_categories(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.delete()
    parts = callback.data.split("_")
    category_id = int(parts[2])
    page = int(parts[4])
    await send_products_page(callback.message, page, state, category_id=category_id)
    await callback.answer()

@router.callback_query(F.data.startswith("show_fav_"))
async def navigate_favorites(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.delete()
    page = int(callback.data.split("_")[3])
    await send_products_page(callback.message, page, state, from_favorites=True)
    await callback.answer()

@router.callback_query(F.data.startswith("select_char_"))
async def show_characteristic_selector(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[2])
    await state.update_data(current_product_id=product_id) # Сохраняем ID товара для возврата

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT id, name, value, price, quantity FROM product_characteristics WHERE product_id = ? ORDER BY price ASC", (product_id,))
        characteristics = await cursor.fetchall()
        
        cursor_product = await db.execute("SELECT name, description, image_url FROM products WHERE id = ?", (product_id,))
        product_name, product_description, product_image_url = await cursor_product.fetchone()

    if not characteristics:
        await callback.answer("У этого товара нет доступных характеристик.", show_alert=True)
        # Вернуться на предыдущую страницу или обновить карточку товара
        is_fav = await is_product_in_favorites(callback.from_user.id, product_id)
        markup = get_product_card_markup(product_id, False, is_fav) # False, так как нет характеристик
        await callback.message.edit_reply_markup(reply_markup=markup)
        return
    
    # Редактируем сообщение с карточкой товара, чтобы показать выбор характеристик
    caption = f"**{product_name}**\n\n{product_description or 'Описание отсутствует'}\n\nВыберите вариант:"
    
    markup = get_characteristic_selection_markup(product_id, characteristics)
    
    # Если это фото, редактируем подпись и кнопки
    if callback.message.photo:
        await callback.message.edit_caption(caption=caption, reply_markup=markup, parse_mode="Markdown")
    else: # Если это текстовое сообщение
        await callback.message.edit_text(text=caption, reply_markup=markup, parse_mode="Markdown")
        
    await callback.answer()

@router.callback_query(F.data.startswith("buy_char_"))
async def show_quantity_selector_for_characteristic(callback: types.CallbackQuery, state: FSMContext):
    characteristic_id = int(callback.data.split("_")[2])
    await state.update_data(current_characteristic_id=characteristic_id) # Сохраняем ID характеристики

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name, value, price, quantity FROM product_characteristics WHERE id = ?", (characteristic_id,))
        char_name, char_value, char_price, char_qty = await cursor.fetchone()
        
        
        # Получаем ID продукта, к которому относится характеристика
        cursor_product_id = await db.execute("SELECT product_id FROM product_characteristics WHERE id = ?", (characteristic_id,))
        product_id = (await cursor_product_id.fetchone())[0]
        await state.update_data(current_product_id=product_id) # Обновляем product_id в состоянии

    if not char_qty or char_qty == 0:
        await callback.answer("Этого варианта нет в наличии!", show_alert=True)
        return
    
    # --- ИСПРАВЛЕННЫЙ КОД НАЧИНАЕТСЯ ЗДЕСЬ ---
    # 1. Инициализируем базовое содержимое подписи/текста пустым значением
    base_caption_content = ""
    
    # 2. Безопасно получаем базовый текст из callback.message
    # Проверяем, есть ли у сообщения подпись (для фото/видео)
    if callback.message.caption:
        # Разбиваем подпись по "Выберите вариант:" и берем часть до него
        parts = callback.message.caption.split('Выберите вариант:', 1)
        base_caption_content = parts[0].strip() # .strip() убирает лишние пробелы
    elif callback.message.text:
        # Если это обычное текстовое сообщение
        # Разбиваем текст по "Выберите вариант:" и берем часть до него
        parts = callback.message.text.split('Выберите вариант:', 1)
        base_caption_content = parts[0].strip()

    # 3. Формируем полную подпись сообщения.
    # Эта часть теперь гарантированно выполняется, если функция не вернулась выше.
    caption = (
        f"{base_caption_content}\n\n"
        f"Выбран: **{char_name}: {char_value}**\n"
        f"Цена: {char_price / 100:.2f}₽\n"
        f"В наличии: {char_qty} шт.\n\n"
        f"Выберите количество:"
    )
    # --- ИСПРАВЛЕННЫЙ КОД ЗАВЕРШАЕТСЯ ЗДЕСЬ ---
    
    markup = get_quantity_selector_markup(characteristic_id, 1, char_qty)

    # Используем edit_caption для фото и edit_text для обычных сообщений
    if callback.message.photo:
        await callback.message.edit_caption(caption=caption, reply_markup=markup, parse_mode="Markdown")
    else:
        await callback.message.edit_text(text=caption, reply_markup=markup, parse_mode="Markdown")
        
    await callback.answer()

@router.callback_query(F.data.startswith("noop_char_no_stock"))
async def noop_char_no_stock_callback(callback: types.CallbackQuery):
    await callback.answer("Этого варианта нет в наличии.", show_alert=True)

@router.callback_query(F.data.startswith("back_to_product_view_"))
async def back_to_product_view_callback(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[4])
    await state.update_data(current_product_id=product_id) # Убедимся, что product_id в состоянии
    
    async with aiosqlite.connect("products.db") as db:
        cursor_product = await db.execute("SELECT name, description, image_url FROM products WHERE id = ?", (product_id,))
        product_name, product_description, product_image_url = await cursor_product.fetchone()
        
        cursor_char_count = await db.execute("SELECT COUNT(*) FROM product_characteristics WHERE product_id = ?", (product_id,))
        has_characteristics = (await cursor_char_count.fetchone())[0] > 0

    caption = f"**{product_name}**\n\n{product_description or 'Описание отсутствует'}"
    is_fav = await is_product_in_favorites(callback.from_user.id, product_id)
    markup = get_product_card_markup(product_id, has_characteristics, is_fav)

    if not has_characteristics:
        caption += "\n\n**Нет доступных характеристик**"
        markup = None # Убираем кнопки, если нельзя купить
    
    if callback.message.photo:
        await callback.message.edit_caption(caption=caption, reply_markup=markup, parse_mode="Markdown")
    else:
        await callback.message.edit_text(text=caption, reply_markup=markup, parse_mode="Markdown")
    await callback.answer()

@router.callback_query(F.data.startswith("cancel_qty_selector_char_"))
async def cancel_quantity_selector_char(callback: types.CallbackQuery, state: FSMContext):
    try:
        # Изменяем разбор callback данных
        characteristic_id = int(callback.data.split("cancel_qty_selector_char_")[1])
        
        async with aiosqlite.connect("products.db") as db:
            # Получаем product_id и информацию о характеристиках
            cursor = await db.execute("SELECT product_id FROM product_characteristics WHERE id = ?", (characteristic_id,))
            product_id = (await cursor.fetchone())[0]

            cursor_chars = await db.execute(
                "SELECT id, name, value, price, quantity FROM product_characteristics WHERE product_id = ? ORDER BY price ASC", 
                (product_id,)
            )
            characteristics = await cursor_chars.fetchall()

            # Получаем информацию о товаре
            cursor_product = await db.execute(
                "SELECT name, description, image_url FROM products WHERE id = ?", 
                (product_id,)
            )
            product_name, product_description, product_image_url = await cursor_product.fetchone()

        caption = f"**{product_name}**\n\n{product_description or 'Описание отсутствует'}\n\nВыберите вариант:"
        markup = get_characteristic_selection_markup(product_id, characteristics)

        if callback.message.photo:
            await callback.message.edit_caption(
                caption=caption,
                reply_markup=markup,
                parse_mode="Markdown"
            )
        else:
            await callback.message.edit_text(
                text=caption,
                reply_markup=markup,
                parse_mode="Markdown"
            )
        
        await callback.answer()
    
    except Exception as e:
        print(f"DEBUG: Ошибка в cancel_quantity_selector_char: {e}")
        await callback.answer("Произошла ошибка при возврате к выбору характеристик.", show_alert=True)


@router.callback_query(F.data.startswith("qty-char-change_"))
async def change_quantity_for_characteristic(callback: types.CallbackQuery, state: FSMContext):
    try:
        _, action, characteristic_id_str, current_qty_str = callback.data.split("_")
        characteristic_id = int(characteristic_id_str)
        current_qty = int(current_qty_str)
        
        async with aiosqlite.connect("products.db") as db:
            cursor = await db.execute("SELECT name, value, price, quantity FROM product_characteristics WHERE id = ?", (characteristic_id,))
            result = await cursor.fetchone()
            
            if not result:
                await callback.answer("Извините, этот вариант товара больше недоступен.", show_alert=True)
                # Обновить сообщение, чтобы убрать кнопки покупки
                if callback.message.photo:
                    await callback.message.edit_caption(
                        caption=f"{callback.message.caption.split('Выберите количество:')[0]}\n\n**ВАРИАНТ БОЛЬШЕ НЕДОСТУПЕН**",
                        parse_mode="Markdown"
                    )
                else:
                    await callback.message.edit_text(
                        text=f"{callback.message.text.split('Выберите количество:')[0]}\n\n**ВАРИАНТ БОЛЬШЕ НЕДОСТУПЕН**",
                        parse_mode="Markdown"
                    )
                return
            
            char_name, char_value, char_price, max_qty = result
            new_qty = current_qty

            if action == "inc":
                if current_qty < max_qty:
                    new_qty = current_qty + 1
                else:
                    await callback.answer(f"В наличии только {max_qty} шт.", show_alert=True)
            elif action == "dec":
                if current_qty > 1:
                    new_qty = current_qty - 1
            
            if new_qty != current_qty:
                markup = get_quantity_selector_markup(characteristic_id, new_qty, max_qty)
                base_caption_content = ""
                if callback.message.caption:
                    parts = callback.message.caption.split('Выберите количество:', 1)
                    base_caption_content = parts[0].strip()
                elif callback.message.text:
                    parts = callback.message.text.split('Выберите количество:', 1)
                    base_caption_content = parts[0].strip()

                updated_caption = (
                    f"{base_caption_content}\n"
                    f"Выберите количество:"
                )
                
                if callback.message.photo:
                    await callback.message.edit_caption(caption=updated_caption, reply_markup=markup, parse_mode="Markdown")
                else:
                    await callback.message.edit_text(text=updated_caption, reply_markup=markup, parse_mode="Markdown")
            
        await callback.answer()
    except Exception as e:
        print(f"DEBUG: Ошибка в change_quantity_for_characteristic: {e}")
        await callback.answer("Ошибка при изменении количества.", show_alert=True)


@router.callback_query(F.data.startswith("cart_add_char_"))
async def add_characteristic_to_cart(callback: types.CallbackQuery, state: FSMContext):
    try:
        # Разбираем callback данные
        parts = callback.data.split("_")
        characteristic_id = int(parts[3])  # cart_add_char_ID_QTY
        quantity_to_add = int(parts[4])
        user_id = callback.from_user.id

        print(f"DEBUG: Adding to cart - char_id: {characteristic_id}, qty: {quantity_to_add}, user_id: {user_id}")

        async with aiosqlite.connect("products.db") as db:
            # Проверяем наличие характеристики и получаем информацию о ней
            cursor = await db.execute("""
                SELECT pc.quantity, pc.name, pc.value, pc.price, p.name as product_name
                FROM product_characteristics pc
                JOIN products p ON pc.product_id = p.id
                WHERE pc.id = ?
            """, (characteristic_id,))
            result = await cursor.fetchone()
            
            if not result:
                print(f"DEBUG: Characteristic {characteristic_id} not found")
                await callback.answer("Этот вариант товара больше недоступен.", show_alert=True)
                return
            
            available_qty, char_name, char_value, char_price, product_name = result
            print(f"DEBUG: Found characteristic - {char_name}: {char_value}, available: {available_qty}")

            # Загружаем текущую корзину пользователя
            cart = await load_user_cart_from_db(user_id)
            current_qty_in_cart = 0
            item_index = -1

            # Проверяем, есть ли уже этот товар в корзине
            for i, (cid, qty) in enumerate(cart):
                if cid == characteristic_id:
                    current_qty_in_cart = qty
                    item_index = i
                    break

            # Проверяем доступное количество
            if current_qty_in_cart + quantity_to_add > available_qty:
                print(f"DEBUG: Not enough quantity. Available: {available_qty}, In cart: {current_qty_in_cart}, Trying to add: {quantity_to_add}")
                await callback.answer(f"Недостаточно товара. Доступно: {available_qty - current_qty_in_cart} шт.", show_alert=True)
                return

            # Обновляем корзину
            if item_index != -1:
                cart[item_index] = (characteristic_id, current_qty_in_cart + quantity_to_add)
            else:
                cart.append((characteristic_id, quantity_to_add))
            
            # Сохраняем обновленную корзину
            await save_user_cart_to_db(user_id, cart)
            print(f"DEBUG: Cart updated successfully for user {user_id}")

            # Обновляем сообщение с выбором количества
            markup = get_quantity_selector_markup(characteristic_id, quantity_to_add + current_qty_in_cart, available_qty)
            
            # Формируем обновленный текст
            base_caption_content = ""
            if callback.message.caption:
                parts = callback.message.caption.split('Выберите количество:', 1)
                base_caption_content = parts[0].strip()
            elif callback.message.text:
                parts = callback.message.text.split('Выберите количество:', 1)
                base_caption_content = parts[0].strip()

            updated_caption = (
                f"{base_caption_content}\n"
                f"Количество в корзине: {quantity_to_add + current_qty_in_cart} шт.\n"
                f"Выберите количество:"
            )

            # Обновляем сообщение
            if callback.message.photo:
                await callback.message.edit_caption(
                    caption=updated_caption,
                    reply_markup=markup,
                    parse_mode="Markdown"
                )
            else:
                await callback.message.edit_text(
                    text=updated_caption,
                    reply_markup=markup,
                    parse_mode="Markdown"
                )

            await callback.answer(
                f"Добавлено в корзину: {quantity_to_add} шт. ({char_name}: {char_value})",
                show_alert=True
            )

    except Exception as e:
        print(f"ERROR in add_characteristic_to_cart: {str(e)}")
        await callback.answer("Произошла ошибка при добавлении в корзину.", show_alert=True)


# --- Логика корзины (обновление) ---
@router.message(F.text == "🛒 Корзина")
async def show_cart(message: types.Message):
    user_id = message.from_user.id
    cart = await load_user_cart_from_db(user_id)

    if not cart:
        await message.answer("Ваша корзина пуста.")
        return

    cart_text = "**Ваша корзина:**\n\n"
    total_price = 0
    item_details = []

    async with aiosqlite.connect("products.db") as db:
        for characteristic_id, quantity in cart:
            cursor = await db.execute("""
                SELECT pc.name, pc.value, pc.price, p.name as product_name
                FROM product_characteristics pc
                JOIN products p ON pc.product_id = p.id
                WHERE pc.id = ?""", (characteristic_id,))
            result = await cursor.fetchone()
            
            if result:
                char_name, char_value, char_price, product_name = result
                item_price = (char_price * quantity) / 100
                total_price += item_price
                item_details.append(f"• {product_name} ({char_name}: {char_value}): {quantity} шт. x {char_price / 100:.2f}₽ = {item_price:.2f}₽")
            else:
                item_details.append(f"• Неизвестный товар (ID характеристики: {characteristic_id}): {quantity} шт. (недоступно)")
    
    cart_text += "\n".join(item_details)
    cart_text += f"\n\n**Итого: {total_price:.2f}₽**"

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Оформить заказ", callback_data="checkout")],
        [InlineKeyboardButton(text="🗑️ Очистить корзину", callback_data="clear_cart")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_main_menu")]
    ])
    await message.answer(cart_text, reply_markup=markup, parse_mode="Markdown")

@router.callback_query(F.data == "clear_cart")
async def clear_cart_callback(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    await save_user_cart_to_db(user_id, []) # Очищаем корзину
    await callback.message.edit_text("Ваша корзина очищена.")
    await callback.answer("Корзина очищена", show_alert=True)

# --- Логика оформления заказа ---
@router.callback_query(F.data == "checkout")
async def checkout_start(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    cart = await load_user_cart_from_db(user_id)
    if not cart:
        await callback.answer("Ваша корзина пуста, невозможно оформить заказ.", show_alert=True)
        await callback.message.delete()
        await show_cart(callback.message)
        return

    await state.set_state(UserStates.awaiting_delivery_name)
    await callback.message.edit_text("Пожалуйста, введите ваше ФИО для доставки:")
    await callback.answer()

@router.message(UserStates.awaiting_delivery_name)
async def process_delivery_name(message: types.Message, state: FSMContext):
    await state.update_data(delivery_name=message.text)
    await state.set_state(UserStates.awaiting_delivery_address)
    await message.answer("Введите адрес доставки (индекс, город, улица, дом, квартира):")

@router.message(UserStates.awaiting_delivery_address)
async def process_delivery_address(message: types.Message, state: FSMContext):
    await state.update_data(delivery_address=message.text)
    await state.set_state(UserStates.awaiting_delivery_phone)
    await message.answer("Введите ваш контактный телефон:")

@router.message(UserStates.awaiting_delivery_phone)
async def process_delivery_phone(message: types.Message, state: FSMContext):
    await state.update_data(delivery_phone=message.text)
    
    user_data = await state.get_data()
    user_id = message.from_user.id
    cart = await load_user_cart_from_db(user_id)

    order_summary = "**Подтвердите ваш заказ:**\n\n"
    order_summary += f"ФИО: {user_data.get('delivery_name')}\n"
    order_summary += f"Адрес: {user_data.get('delivery_address')}\n"
    order_summary += f"Телефон: {user_data.get('delivery_phone')}\n\n"
    order_summary += "**Состав заказа:**\n"

    total_price = 0
    async with aiosqlite.connect("products.db") as db:
        for characteristic_id, quantity in cart:
            cursor = await db.execute("""
                SELECT pc.name, pc.value, pc.price, p.name as product_name
                FROM product_characteristics pc
                JOIN products p ON pc.product_id = p.id
                WHERE pc.id = ?""", (characteristic_id,))
            result = await cursor.fetchone()
            if result:
                char_name, char_value, char_price, product_name = result
                item_price = (char_price * quantity) / 100
                total_price += item_price
                order_summary += f"• {product_name} ({char_name}: {char_value}): {quantity} шт. x {char_price / 100:.2f}₽ = {item_price:.2f}₽\n"
            else:
                order_summary += f"• Неизвестный товар (ID характеристики: {characteristic_id}): {quantity} шт. (недоступно)\n"
    
    order_summary += f"\n**Итого: {total_price:.2f}₽**"

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Подтвердить", callback_data="confirm_order")],
        [InlineKeyboardButton(text="❌ Отменить", callback_data="cancel_order")]
    ])
    await message.answer(order_summary, reply_markup=markup, parse_mode="Markdown")
    await state.set_state(UserStates.awaiting_order_confirmation)

@router.callback_query(F.data == "confirm_order", StateFilter(UserStates.awaiting_order_confirmation))
async def confirm_order(callback: types.CallbackQuery, state: FSMContext):
    user_data = await state.get_data()
    user_id = callback.from_user.id
    cart = await load_user_cart_from_db(user_id)

    if not cart:
        await callback.answer("Ваша корзина пуста. Заказ уже оформлен или очищен.", show_alert=True)
        await callback.message.delete()
        await start(callback.message)
        await state.clear()
        return

    delivery_info = (
        f"ФИО: {user_data.get('delivery_name')}\n"
        f"Адрес: {user_data.get('delivery_address')}\n"
        f"Телефон: {user_data.get('delivery_phone')}"
    )

    async with aiosqlite.connect("products.db") as db:
        # Создаем новый заказ с датой создания
        cursor = await db.execute(
            "INSERT INTO orders (user_id, delivery_info, status, created_at) VALUES (?, ?, ?, datetime('now'))",
            (user_id, delivery_info, 'Оформлен')
        )
        order_id = cursor.lastrowid
        
        order_items_text = ""
        total_order_price = 0

        # 2. Добавляем товары из корзины в order_items и уменьшаем количество на складе
        for characteristic_id, quantity_in_cart in cart:
            # Получаем текущее количество и информацию о товаре
            cursor_item = await db.execute(
                "SELECT quantity, product_id, name, value, price FROM product_characteristics WHERE id = ?",
                (characteristic_id,)
            )
            item_info = await cursor_item.fetchone()

            if item_info:
                available_qty, product_id, char_name, char_value, char_price = item_info
                
                # Проверяем, достаточно ли товара на складе
                if available_qty >= quantity_in_cart:
                    new_qty_on_stock = available_qty - quantity_in_cart
                    await db.execute(
                        "UPDATE product_characteristics SET quantity = ? WHERE id = ?",
                        (new_qty_on_stock, characteristic_id)
                    )
                    await db.execute("INSERT INTO order_items (order_id, characteristic_id, quantity) VALUES (?, ?, ?)",
                                     (order_id, characteristic_id, quantity_in_cart))
                    
                    product_cursor = await db.execute("SELECT name FROM products WHERE id = ?", (product_id,))
                    result = await product_cursor.fetchone()
                    product_name = result[0] if result else "Неизвестный товар"
                    
                    item_total_price = (char_price * quantity_in_cart) / 100
                    total_order_price += item_total_price
                    order_items_text += f"• {product_name} ({char_name}: {char_value}): {quantity_in_cart} шт. x {char_price / 100:.2f}₽ = {item_total_price:.2f}₽\n"
                else:
                    # Если товара недостаточно, информируем пользователя и не добавляем в заказ
                    await callback.answer(f"Недостаточно товара '{char_name}: {char_value}' на складе. Доступно: {available_qty} шт.", show_alert=True)
                    await state.clear()
                    await start(callback.message)
                    return # Прерываем оформление заказа
            else:
                await callback.answer(f"Товар с характеристикой ID {characteristic_id} не найден. Отмена заказа.", show_alert=True)
                await state.clear()
                await start(callback.message)
                return # Прерываем оформление заказа

        await db.commit()
        await save_user_cart_to_db(user_id, []) # Очищаем корзину после успешного оформления

    admin_notification_text = (
        f"🎉 **НОВЫЙ ЗАКАЗ №{order_id}** 🎉\n\n"
        f"**От пользователя:** @{callback.from_user.username or user_id} (ID: {user_id})\n\n"
        f"**Данные для доставки:**\n{delivery_info}\n\n"
        f"**Состав заказа:**\n{order_items_text}\n"
        f"**Общая сумма: {total_order_price:.2f}₽**"
    )
    await notify_admin(admin_notification_text)

    await callback.message.edit_text("Ваш заказ успешно оформлен! Мы свяжемся с вами в ближайшее время.")
    await state.clear()
    await callback.answer()
    await start(callback.message) # Возвращаем пользователя в главное меню

@router.callback_query(F.data == "cancel_order", StateFilter(UserStates.awaiting_order_confirmation))
async def cancel_order(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("Оформление заказа отменено.")
    await callback.answer()
    await start(callback.message)

# --- Мои заказы ---
@router.message(F.text == "📦 Мои заказы")
async def my_orders(message: types.Message):
    user_id = message.from_user.id
    try:
        async with aiosqlite.connect("products.db") as db:
            cursor = await db.execute("""
                SELECT o.id, o.delivery_info, o.status, 
                       COUNT(oi.id) as items_count,
                       SUM(oi.quantity * pc.price) as total_price
                FROM orders o
                LEFT JOIN order_items oi ON o.id = oi.order_id
                LEFT JOIN product_characteristics pc ON oi.characteristic_id = pc.id
                WHERE o.user_id = ?
                GROUP BY o.id
                ORDER BY o.id DESC
            """, (user_id,))
            orders = await cursor.fetchall()

            if not orders:
                await message.answer("У вас пока нет заказов.")
                return

            # Создаем кнопки для каждого заказа
            buttons = []
            for order in orders:
                order_id, delivery_info, status, items_count, total_price = order
                total_price = total_price / 100 if total_price else 0
                
                button_text = f"Заказ №{order_id} | {status} | {total_price:.2f}₽"
                buttons.append([InlineKeyboardButton(
                    text=button_text,
                    callback_data=f"view_order_{order_id}"
                )])

            markup = InlineKeyboardMarkup(inline_keyboard=buttons)
            await message.answer("Ваши заказы:", reply_markup=markup)

    except Exception as e:
        print(f"Error in my_orders: {e}")
        await message.answer("Произошла ошибка при получении истории заказов. Пожалуйста, попробуйте позже.")


# --- Админ: Управление заказами ---
@router.message(F.text == "📥 Заказы (админ)", F.from_user.id == YOUR_ADMIN_ID)
async def admin_orders(message: types.Message):
    try:
        async with aiosqlite.connect("products.db") as db:
            # Получаем все заказы с детальной информацией
            cursor = await db.execute("""
                SELECT 
                    o.id,
                    o.user_id,
                    o.status,
                    o.delivery_info,
                    COUNT(oi.id) as items_count,
                    SUM(oi.quantity * pc.price) as total_price
                FROM orders o
                LEFT JOIN order_items oi ON o.id = oi.order_id
                LEFT JOIN product_characteristics pc ON oi.characteristic_id = pc.id
                GROUP BY o.id
                ORDER BY o.id DESC
            """)
            orders = await cursor.fetchall()

            if not orders:
                await message.answer("Пока нет заказов.")
                return

            # Создаем кнопки для каждого заказа
            buttons = []
            for order in orders:
                order_id, user_id, status, delivery_info, items_count, total_price = order
                total_price = total_price / 100 if total_price else 0  # Конвертируем копейки в рубли
                
                # Формируем текст кнопки
                button_text = (f"Заказ #{order_id} | {status}\n"
                             f"От: ID {user_id} | {items_count} поз. | {total_price:.2f}₽")
                
                buttons.append([InlineKeyboardButton(
                    text=button_text,
                    callback_data=f"view_order_{order_id}"
                )])

            # Добавляем кнопку возврата в админ-панель
            buttons.append([InlineKeyboardButton(
                text="↩️ Вернуться в админ-панель",
                callback_data="back_to_admin_panel"
            )])

            markup = InlineKeyboardMarkup(inline_keyboard=buttons)
            await message.answer("Список заказов:", reply_markup=markup)

    except Exception as e:
        print(f"Error in admin_orders: {e}")
        await message.answer("Произошла ошибка при получении списка заказов. Пожалуйста, попробуйте позже.")

# Добавьте также обработчик для просмотра деталей заказа
@router.callback_query(F.data.startswith("view_order_"))
async def view_order_details(callback: types.CallbackQuery):
    try:
        order_id = int(callback.data.split("_")[2])
        is_admin = callback.from_user.id == YOUR_ADMIN_ID  # Проверка на админа
        
        async with aiosqlite.connect("products.db") as db:
            # Получаем основную информацию о заказе
            cursor = await db.execute("""
                SELECT o.id, o.user_id, o.status, o.delivery_info, o.created_at
                FROM orders o
                WHERE o.id = ?
            """, (order_id,))
            order_info = await cursor.fetchone()
            
            if not order_info:
                await callback.message.edit_text("Заказ не найден.")
                return

            # Проверяем права доступа (админ может смотреть все заказы, пользователь - только свои)
            if not is_admin and order_info[1] != callback.from_user.id:
                await callback.message.edit_text("У вас нет доступа к этому заказу.")
                return

            # Получаем детали заказа
            cursor = await db.execute("""
                SELECT 
                    oi.quantity,
                    pc.name as char_name,
                    pc.value as char_value,
                    pc.price,
                    p.name as product_name
                FROM order_items oi
                JOIN product_characteristics pc ON oi.characteristic_id = pc.id
                JOIN products p ON pc.product_id = p.id
                WHERE oi.order_id = ?
            """, (order_id,))
            items = await cursor.fetchall()

            # Формируем текст заказа
            order_text = f"Заказ №{order_info[0]}\n"
            order_text += f"От пользователя: ID {order_info[1]}\n"
            order_text += f"Статус: {order_info[2]}\n"
            if order_info[4]:  # если есть дата создания
                order_text += f"Дата создания: {order_info[4]}\n"
            order_text += f"\nИнформация о доставке:\n{order_info[3]}\n\n"
            order_text += "Состав заказа:\n"

            total_sum = 0
            for item in items:
                quantity, char_name, char_value, price, product_name = item
                item_sum = (price * quantity) / 100
                total_sum += item_sum
                
                order_text += f"• {product_name} ({char_name}: {char_value})\n"
                order_text += f"  {quantity} шт. × {price/100:.2f}₽ = {item_sum:.2f}₽\n"

            order_text += f"\nОбщая сумма: {total_sum:.2f}₽"

            # Создаем разные наборы кнопок для админа и пользователя
            markup_buttons = []
            
            if is_admin:
                # Кнопки управления статусом только для админа
                markup_buttons.extend([
                    [InlineKeyboardButton(text="✅ Выполнен", callback_data=f"set_{order_id}_Выполнен"),
                     InlineKeyboardButton(text="❌ Отменён", callback_data=f"set_{order_id}_Отменён")],
                    [InlineKeyboardButton(text="🔄 В обработке", callback_data=f"set_{order_id}_В_обработке")],
                    [InlineKeyboardButton(text="📦 Отправлен", callback_data=f"set_{order_id}_Отправлен")],
                ])
            
            # Кнопка возврата к списку для всех
            markup_buttons.append([InlineKeyboardButton(text="↩️ Назад к списку", 
                                                      callback_data="back_to_orders_list" if is_admin else "back_to_my_orders")])

            markup = InlineKeyboardMarkup(inline_keyboard=markup_buttons)

            await callback.message.edit_text(order_text, reply_markup=markup)
            await callback.answer()
            
    except Exception as e:
        print(f"Error in view_order_details: {e}")
        await callback.message.edit_text("Произошла ошибка при получении деталей заказа.")

@router.callback_query(F.data == "back_to_my_orders")
async def back_to_my_orders(callback: types.CallbackQuery):
    await callback.message.delete()
    await my_orders(callback.message)  # Возвращаемся к списку заказов пользователя
    await callback.answer()

# Добавьте обработчик для возврата к списку заказов
@router.callback_query(F.data == "back_to_orders_list")
async def back_to_orders_list(callback: types.CallbackQuery):
    await callback.message.delete()
    await admin_orders(callback.message)

@router.callback_query(F.data.startswith("set_"))
async def update_order_status(callback: types.CallbackQuery):
    try:
        # Получаем данные из callback_data
        callback_data = callback.data.split("_", 2)  # Разделяем только на 3 части
        if len(callback_data) != 3:
            await callback.answer("Неверный формат данных", show_alert=True)
            return
            
        order_id = int(callback_data[1])
        new_status = callback_data[2]

        async with aiosqlite.connect("products.db") as db:
            # Обновляем статус заказа
            await db.execute(
                "UPDATE orders SET status = ? WHERE id = ?",
                (new_status, order_id)
            )
            await db.commit()

            # Получаем обновленную информацию о заказе
            cursor = await db.execute("""
                SELECT o.id, o.user_id, o.status, o.delivery_info
                FROM orders o
                WHERE o.id = ?
            """, (order_id,))
            order_info = await cursor.fetchone()

            if not order_info:
                await callback.answer("Заказ не найден", show_alert=True)
                return

            # Получаем детали заказа
            cursor = await db.execute("""
                SELECT 
                    oi.quantity,
                    pc.name as char_name,
                    pc.value as char_value,
                    pc.price,
                    p.name as product_name
                FROM order_items oi
                JOIN product_characteristics pc ON oi.characteristic_id = pc.id
                JOIN products p ON pc.product_id = p.id
                WHERE oi.order_id = ?
            """, (order_id,))
            items = await cursor.fetchall()

            # Формируем обновленный текст заказа
            order_text = (f"**Заказ #{order_info[0]}**\n"
                         f"От пользователя: ID {order_info[1]}\n"
                         f"Статус: {order_info[2]}\n\n"
                         f"**Информация о доставке:**\n{order_info[3]}\n\n"
                         f"**Состав заказа:**\n")

            total_sum = 0
            for item in items:
                quantity, char_name, char_value, price, product_name = item
                item_sum = (price * quantity) / 100
                total_sum += item_sum
                order_text += (f"- {product_name} ({char_name}: {char_value})\n"
                             f"  {quantity} шт. × {price/100:.2f}₽ = {item_sum:.2f}₽\n")

            order_text += f"\n**Общая сумма**: {total_sum:.2f}₽"

            # Обновляем кнопки статуса
            markup = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✅ Выполнен", callback_data=f"set_{order_id}_Выполнен"),
                 InlineKeyboardButton(text="❌ Отменён", callback_data=f"set_{order_id}_Отменён")],
                [InlineKeyboardButton(text="🔄 В обработке", callback_data=f"set_{order_id}_В_обработке")],
                [InlineKeyboardButton(text="📦 Отправлен", callback_data=f"set_{order_id}_Отправлен")],
                [InlineKeyboardButton(text="↩️ Назад к списку", callback_data="back_to_orders_list")]
            ])

            # Обновляем сообщение с новым статусом
            await callback.message.edit_text(order_text, reply_markup=markup, parse_mode="Markdown")

            # Отправляем уведомление пользователю
            try:
                await bot.send_message(
                    order_info[1],  # user_id
                    f"Статус вашего заказа #{order_info[0]} изменен на: {new_status}"
                )
            except Exception as e:
                print(f"Failed to notify user {order_info[1]} about status change: {e}")

            await callback.answer(f"Статус заказа изменен на: {new_status}")

    except Exception as e:
        print(f"Error in update_order_status: {e}")
        await callback.answer("Произошла ошибка при обновлении статуса заказа", show_alert=True)


# --- Админ: Экспорт заказов в Excel ---
@router.message(F.text == "📤 Экспорт заказов", F.from_user.id == YOUR_ADMIN_ID)
async def export_orders(message: types.Message):
    try:
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"
        
        # Задаем заголовки
        headers = [
            "ID Заказа", 
            "Дата создания",
            "ID Пользователя", 
            "Информация о доставке", 
            "Статус", 
            "Товар",
            "Характеристика",
            "Количество",
            "Цена за ед. (₽)",
            "Сумма (₽)"
        ]
        ws.append(headers)

        # Подключаемся к базе данных
        async with aiosqlite.connect("products.db") as db:
            try:
                # Получаем данные с помощью LEFT JOIN
                cursor = await db.execute("""
                    SELECT 
                        o.id,
                        o.created_at,
                        o.user_id,
                        o.delivery_info,
                        o.status,
                        p.name as product_name,
                        pc.name as char_name,
                        pc.value as char_value,
                        pc.price,
                        oi.quantity
                    FROM orders o
                    LEFT JOIN order_items oi ON o.id = oi.order_id
                    LEFT JOIN product_characteristics pc ON oi.characteristic_id = pc.id
                    LEFT JOIN products p ON pc.product_id = p.id
                    ORDER BY o.id DESC
                """)
                
                orders = await cursor.fetchall()
                print(f"DEBUG: Retrieved {len(orders)} orders from database")

                if not orders:
                    await message.answer("Нет заказов для экспорта.")
                    return

                # Добавляем данные в Excel
                for order in orders:
                    order_id = order[0]
                    created_at = order[1] or "Не указано"
                    user_id = order[2]
                    delivery_info = order[3]
                    status = order[4]
                    product_name = order[5] or "Товар удален"
                    char_name = order[6]
                    char_value = order[7]
                    price = order[8]
                    quantity = order[9]

                    # Вычисляем цену и общую сумму
                    if price is not None and quantity is not None:
                        price_rub = price / 100
                        total = (price * quantity) / 100
                    else:
                        price_rub = 0
                        total = 0

                    # Формируем характеристику
                    characteristic = f"{char_name}: {char_value}" if char_name and char_value else "Н/Д"

                    # Добавляем строку в Excel
                    ws.append([
                        order_id,
                        created_at,
                        user_id,
                        delivery_info,
                        status,
                        product_name,
                        characteristic,
                        quantity or 0,
                        price_rub,
                        total
                    ])

                # Форматирование
                for cell in ws[1]:
                    cell.font = Font(bold=True)

                # Автоматическая настройка ширины столбцов
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

                # Сохраняем файл
                current_time = datetime.now().strftime("%d%m%Y")
                filename = f"orders_export_{current_time}.xlsx"
                wb.save(filename)

                # Отправляем файл
                try:
                    await message.answer_document(
                        document=FSInputFile(filename),
                        caption=f"Экспорт заказов завершен успешно!\nВсего заказов: {len(orders)}\nДата экспорта: {datetime.now().strftime('%d-%m-%Y')}"
                    )
                finally:
                    # Удаляем временный файл
                    if os.path.exists(filename):
                        os.remove(filename)

            except Exception as db_error:
                print(f"Database error: {db_error}")
                raise

    except Exception as e:
        print(f"Error in export_orders: {str(e)}")
        await message.answer(
            "Произошла ошибка при экспорте заказов. "
            "Пожалуйста, попробуйте позже или обратитесь к администратору."
        )
    
# Измените обработчик, чтобы он соответствовал тексту кнопки
@router.message(F.text == "📤 Экспорт товаров", F.from_user.id == YOUR_ADMIN_ID)  # Изменено с "📊" на "📤"
async def export_products(message: types.Message):
    print("DEBUG: export_products function called")  # Добавьте эту строку
    try:
        async with aiosqlite.connect("products.db") as db:
            # Получаем все товары с их характеристиками
            cursor = await db.execute("""
                SELECT 
                    p.id as product_id,
                    p.name as product_name,
                    p.description,
                    pc.name as char_name,
                    pc.value as char_value,
                    pc.price,
                    pc.quantity
                FROM products p
                LEFT JOIN product_characteristics pc ON p.id = pc.product_id
                ORDER BY p.id, pc.id
            """)
            products = await cursor.fetchall()

            if not products:
                await message.answer("В базе данных нет товаров.")
                return

            # Формируем текст отчета
            report_text = "Список товаров:\n\n"
            current_product_id = None
            total_items = 0
            total_value = 0

            for product in products:
                product_id, product_name, description, char_name, char_value, price, quantity = product
                
                if current_product_id != product_id:
                    current_product_id = product_id
                    report_text += f"Товар: {product_name}\n"
                    if description:
                        report_text += f"Описание: {description}\n"
                
                if char_name and char_value is not None:
                    price_rub = price / 100 if price else 0
                    report_text += f"- {char_name}: {char_value}\n"
                    report_text += f"  Цена: {price_rub:.2f}₽\n"
                    report_text += f"  В наличии: {quantity} шт.\n"
                    
                    total_items += quantity if quantity else 0
                    total_value += (price_rub * quantity) if price and quantity else 0
                
                report_text += "\n"

            report_text += f"\nОбщая статистика:\n"
            report_text += f"Всего единиц товара: {total_items}\n"
            report_text += f"Общая стоимость: {total_value:.2f}₽"

            # Создаем временный файл отчета
            current_time = datetime.now().strftime("%d%m%Y")
            filename = f"export_products_{current_time}.txt"
            
            with open(filename, "w", encoding="utf-8") as file:
                file.write(report_text)

            # Отправляем файл
            try:
                with open(filename, "rb") as file:
                    await message.answer_document(
                        document=types.BufferedInputFile(
                            file.read(),
                            filename=filename
                        ),
                        caption="Отчет по товарам"
                    )
            finally:
                # Удаляем временный файл
                if os.path.exists(filename):
                    os.remove(filename)

            # Отправляем краткую сводку
            summary = (
                f"Экспорт завершен\n"
                f"Всего позиций: {len(products)}\n"
                f"Общее количество: {total_items} шт.\n"
                f"Общая стоимость: {total_value:.2f}₽"
            )
            
            await message.answer(summary)

    except Exception as e:
        print(f"Error in export_products: {e}")
        await message.answer("Произошла ошибка при экспорте товаров. Пожалуйста, попробуйте позже.")

@router.message(F.text == "📥 Импорт товаров", F.from_user.id == YOUR_ADMIN_ID)
async def import_products_start(message: types.Message):
    """Начало процесса импорта товаров"""
    instruction_text = """
📝 **Инструкция по импорту товаров:**

1. Подготовьте Excel-файл со следующими столбцами:
   - Название товара
   - Описание
   - Категория
   - Характеристика (название)
   - Характеристика (значение)
   - Цена (в рублях)
   - Количество

2. Пример содержимого: Название | Описание | Категория | Хар-ка (имя) | Хар-ка (знач) | Цена | Кол-во Духи X | Описание | Женский | Объем | 50мл | 1500 | 10

3. Отправьте подготовленный Excel-файл в чат.

❗️ Важно:
- Категория должна существовать в системе
- Цены указывайте в рублях (не копейках)
- Один товар может иметь несколько характеристик
"""
    await message.answer(instruction_text, parse_mode="Markdown")

@router.message(F.document, F.from_user.id == YOUR_ADMIN_ID)
async def process_import_file(message: types.Message):
    """Обработка загруженного файла импорта"""
    if not message.document.file_name.endswith(('.xlsx', '.xls')):
        await message.answer("❌ Пожалуйста, отправьте файл Excel (.xlsx или .xls)")
        return

    try:
        # Скачиваем файл
        file = await bot.get_file(message.document.file_id)
        file_path = file.file_path
        downloaded_file = await bot.download_file(file_path)
        
        # Сохраняем временно
        temp_file_name = f"temp_import_{message.from_user.id}.xlsx"
        with open(temp_file_name, 'wb') as f:
            f.write(downloaded_file.read())

        # Читаем Excel-файл
        wb = load_workbook(temp_file_name)
        ws = wb.active

        # Статистика импорта
        stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'products': {}
        }

        async with aiosqlite.connect("products.db") as db:
            # Получаем существующие категории
            cursor = await db.execute("SELECT name, id FROM categories")
            categories = dict(await cursor.fetchall())

            # Обрабатываем строки файла
            current_product = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Пропускаем пустые строки
                    continue

                name, description, category, char_name, char_value, price, quantity = row
                stats['total'] += 1

                try:
                    # Проверяем/создаем товар
                    if name:
                        category_id = categories.get(category)
                        if not category_id:
                            await message.answer(f"⚠️ Категория '{category}' не найдена для товара '{name}'")
                            stats['failed'] += 1
                            continue

                        cursor = await db.execute(
                            "INSERT INTO products (name, description, category_id) VALUES (?, ?, ?) RETURNING id",
                            (name, description, category_id)
                        )
                        current_product = (await cursor.fetchone())[0]
                        stats['products'][current_product] = name

                    # Добавляем характеристику
                    if current_product and char_name and char_value:
                        price_kopecks = int(float(price) * 100)
                        await db.execute(
                            """INSERT INTO product_characteristics 
                               (product_id, name, value, price, quantity) 
                               VALUES (?, ?, ?, ?, ?)""",
                            (current_product, char_name, char_value, price_kopecks, quantity)
                        )
                        stats['success'] += 1

                except Exception as e:
                    print(f"Error importing row: {e}")
                    stats['failed'] += 1

            await db.commit()

        # Формируем отчет
        report = f"""
📊 **Результаты импорта:**
Всего обработано: {stats['total']}
✅ Успешно: {stats['success']}
❌ С ошибками: {stats['failed']}

**Импортированные товары:**
"""
        for prod_id, prod_name in stats['products'].items():
            report += f"- {prod_name}\n"

        await message.answer(report, parse_mode="Markdown")

    except Exception as e:
        await message.answer(f"❌ Ошибка при импорте: {str(e)}")
    finally:
        # Удаляем временный файл
        if os.path.exists(temp_file_name):
            os.remove(temp_file_name)




# --- Админ: Добавление товара ---
@router.message(F.text == "➕ Добавить товар", F.from_user.id == YOUR_ADMIN_ID)
async def add_product_start(message: types.Message, state: FSMContext):
    await state.clear() # Очищаем предыдущие состояния
    await message.answer("Введите название товара:")
    await state.set_state(AdminStates.add_product_name)

@router.message(AdminStates.add_product_name)
async def process_add_product_name(message: types.Message, state: FSMContext):
    await state.update_data(new_product_name=message.text)
    await message.answer("Введите описание товара:")
    await state.set_state(AdminStates.add_product_description)

@router.message(AdminStates.add_product_description)
async def process_add_product_description(message: types.Message, state: FSMContext):
    await state.update_data(new_product_description=message.text)
    await message.answer("Загрузите изображение товара или введите URL изображения (пропустить: /skip):")
    await state.set_state(AdminStates.add_product_image)

@router.message(AdminStates.add_product_image)
async def process_add_product_image(message: types.Message, state: FSMContext):
    image_url = None
    if message.photo:
        image_url = message.photo[-1].file_id # Используем file_id как URL для Telegram-фото
    elif message.text and message.text.lower() != "/skip":
        image_url = message.text # Предполагаем, что это URL
    
    await state.update_data(new_product_image_url=image_url)

    # Предлагаем выбрать категорию
    markup = await get_categories_markup(include_none=True)
    await message.answer("Выберите категорию для товара:", reply_markup=markup)
    await state.set_state(AdminStates.add_product_category)

@router.callback_query(F.data.startswith("select_category_for_product_"), AdminStates.add_product_category)
async def process_add_product_category_callback(callback: types.CallbackQuery, state: FSMContext):
    category_id = callback.data.split("_")[4]
    if category_id == "none":
        category_id = None
        category_name = "Без категории"
    else:
        category_id = int(category_id)
        category_name = await get_category_name(category_id)

    product_data = await state.get_data()
    product_name = product_data.get('new_product_name')
    product_description = product_data.get('new_product_description')
    product_image_url = product_data.get('new_product_image_url')

    async with aiosqlite.connect("products.db") as db:
        await db.execute(
            "INSERT INTO products (name, description, category_id, image_url) VALUES (?, ?, ?, ?)",
            (product_name, product_description, category_id, product_image_url)
        )
        await db.commit()
    
    await state.clear()
    await callback.message.edit_text(f"Товар '{product_name}' добавлен в категорию '{category_name}'!")
    await callback.answer()


# --- Админ: Управление характеристиками (НОВЫЙ БЛОК) ---

@router.message(F.text == "🎛️ Упр. характеристиками", F.from_user.id == YOUR_ADMIN_ID)
async def manage_characteristics_start(message: types.Message, state: FSMContext):
    await state.clear()
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT id, name FROM products ORDER BY name ASC")
        products = await cursor.fetchall()
    
    if not products:
        await message.answer("Пока нет товаров для управления характеристиками.")
        return

    markup_buttons = []
    for prod_id, prod_name in products:
        markup_buttons.append([InlineKeyboardButton(text=prod_name, callback_data=f"manage_char_for_prod_{prod_id}")])
    
    markup_buttons.append([InlineKeyboardButton(text="↩️ Админ-панель", callback_data="back_to_admin_menu")])
    
    await message.answer("Выберите товар для управления характеристиками:", reply_markup=InlineKeyboardMarkup(inline_keyboard=markup_buttons))


@router.callback_query(F.data.startswith("manage_char_for_prod_"), F.from_user.id == YOUR_ADMIN_ID)
async def show_product_characteristics(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[4])
    await state.update_data(current_product_id_char_manage=product_id)
    
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name FROM products WHERE id = ?", (product_id,))
        product_name = (await cursor.fetchone())[0]

        cursor = await db.execute("SELECT id, name, value, price, quantity FROM product_characteristics WHERE product_id = ?", (product_id,))
        characteristics = await cursor.fetchall()
    
    response_text = f"**Характеристики товара '{product_name}':**\n\n"
    markup_buttons = []

    if characteristics:
        for char_id, name, value, price, quantity in characteristics:
            response_text += f"ID: {char_id} | {name}: {value} | Цена: {price / 100:.2f}₽ | Наличие: {quantity} шт.\n"
            markup_buttons.append([InlineKeyboardButton(text=f"✏️ Изменить ID {char_id}", callback_data=f"edit_char_{char_id}")])
            markup_buttons.append([InlineKeyboardButton(text=f"🗑️ Удалить ID {char_id}", callback_data=f"delete_char_{char_id}")])
        response_text += "\n"
    else:
        response_text += "У этого товара пока нет характеристик.\n\n"

    markup_buttons.append([InlineKeyboardButton(text="➕ Добавить характеристику", callback_data=f"add_char_for_prod_{product_id}")])
    markup_buttons.append([InlineKeyboardButton(text="↩️ Назад к выбору товаров", callback_data="back_to_char_product_list")])
    
    await callback.message.edit_text(response_text, reply_markup=InlineKeyboardMarkup(inline_keyboard=markup_buttons), parse_mode="Markdown")
    await callback.answer()

@router.callback_query(F.data.startswith("add_char_for_prod_"), F.from_user.id == YOUR_ADMIN_ID)
async def add_characteristic_start(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[4])
    await state.update_data(current_product_id_char_manage=product_id) # Убедимся, что product_id в состоянии
    await state.set_state(AdminStates.add_characteristic_name)
    await callback.message.edit_text("Введите название характеристики (например, 'Объем', 'Цвет'):")
    await callback.answer()

@router.message(AdminStates.add_characteristic_name)
async def process_add_characteristic_name(message: types.Message, state: FSMContext):
    await state.update_data(new_char_name=message.text)
    await state.set_state(AdminStates.add_characteristic_value)
    await message.answer("Введите значение характеристики (например, '50мл', 'Красный'):")

@router.message(AdminStates.add_characteristic_value)
async def process_add_characteristic_value(message: types.Message, state: FSMContext):
    await state.update_data(new_char_value=message.text)
    await state.set_state(AdminStates.add_characteristic_price)
    await message.answer("Введите цену для этой характеристики (в копейках, например, 12000 для 120.00₽):")

@router.message(AdminStates.add_characteristic_price)
async def process_add_characteristic_price(message: types.Message, state: FSMContext):
    try:
        price = int(message.text)
        if price <= 0:
            raise ValueError
        await state.update_data(new_char_price=price)
        await state.set_state(AdminStates.add_characteristic_quantity)
        await message.answer("Введите количество этого варианта на складе:")
    except ValueError:
        await message.answer("Некорректная цена. Пожалуйста, введите целое положительное число (в копейках).")

@router.message(AdminStates.add_characteristic_quantity)
async def process_add_characteristic_quantity(message: types.Message, state: FSMContext):
    try:
        quantity = int(message.text)
        if quantity < 0:
            raise ValueError
        await state.update_data(new_char_quantity=quantity)
        
        data = await state.get_data()
        product_id = data.get('current_product_id_char_manage')
        char_name = data.get('new_char_name')
        char_value = data.get('new_char_value')
        char_price = data.get('new_char_price')
        char_quantity = data.get('new_char_quantity')

        async with aiosqlite.connect("products.db") as db:
            await db.execute(
                "INSERT INTO product_characteristics (product_id, name, value, price, quantity) VALUES (?, ?, ?, ?, ?)",
                (product_id, char_name, char_value, char_price, char_quantity)
            )
            await db.commit()
        
        await state.clear()
        await message.answer(f"Характеристика '{char_name}: {char_value}' для товара добавлена!")
        # Возвращаемся к списку характеристик для этого товара
        # Имитируем callback_query для обновления списка
        temp_callback = types.CallbackQuery(id='temp_id', from_user=message.from_user, chat_instance='temp_chat', data=f"manage_char_for_prod_{product_id}", message=message)
        await show_product_characteristics(temp_callback, state)

    except ValueError:
        await message.answer("Некорректное количество. Пожалуйста, введите целое неотрицательное число.")


@router.callback_query(F.data.startswith("edit_char_"), F.from_user.id == YOUR_ADMIN_ID)
async def edit_characteristic_select_field(callback: types.CallbackQuery, state: FSMContext):
    characteristic_id = int(callback.data.split("_")[2])
    await state.update_data(char_to_edit_id=characteristic_id)
    
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name, value, price, quantity FROM product_characteristics WHERE id = ?", (characteristic_id,))
        char_name, char_value, char_price, char_quantity = await cursor.fetchone()

    
    # Безопасное построение кнопки "↩️ Назад" — пытаемся найти callback в последних рядах inline-клавиатуры,
    # если не удаётся — используем явный fallback 'back_to_char_product_list'.
    back_cb = 'back_to_char_product_list'
    try:
        inline_kb = callback.message.reply_markup.inline_keyboard if callback.message.reply_markup else []
        if inline_kb:
            # проверяем последний ряд
            last_row = inline_kb[-1] if len(inline_kb) >= 1 else None
            if last_row and len(last_row) > 0:
                last_cb = getattr(last_row[0], 'callback_data', None)
                if last_cb and 'manage_char_for_prod_' in last_cb:
                    parts = last_cb.split('_')
                    if len(parts) > 4:
                        back_cb = f"manage_char_for_prod_{parts[4]}"
            # проверяем предпоследний ряд
            if back_cb == 'back_to_char_product_list' and len(inline_kb) > 1:
                sec_row = inline_kb[-2]
                if sec_row and len(sec_row) > 0:
                    sec_cb = getattr(sec_row[0], 'callback_data', None)
                    if sec_cb and 'manage_char_for_prod_' in sec_cb:
                        parts = sec_cb.split('_')
                        if len(parts) > 4:
                            back_cb = f"manage_char_for_prod_{parts[4]}"
    except Exception as e:
        print(f"DEBUG: unable to determine back callback: {e}")

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"Название ({char_name})", callback_data="edit_char_field_name")],
        [InlineKeyboardButton(text=f"Значение ({char_value})", callback_data="edit_char_field_value")],
        [InlineKeyboardButton(text=f"Цена ({char_price / 100:.2f}₽)", callback_data="edit_char_field_price")],
        [InlineKeyboardButton(text=f"Количество ({char_quantity} шт.)", callback_data="edit_char_field_quantity")],
        [InlineKeyboardButton(text="↩️ Назад", callback_data=back_cb)]
    ])

    await callback.message.edit_text(f"Выберите поле для редактирования характеристики ID {characteristic_id}:", reply_markup=markup)
    await state.set_state(AdminStates.edit_characteristic_field)
    await callback.answer()


@router.callback_query(F.data.startswith("edit_char_field_"), AdminStates.edit_characteristic_field)
async def edit_characteristic_field_prompt(callback: types.CallbackQuery, state: FSMContext):
    field_to_edit = callback.data.split("_")[3]
    await state.update_data(field_to_edit=field_to_edit)
    
    prompt_text = ""
    if field_to_edit == "name":
        prompt_text = "Введите новое название характеристики:"
    elif field_to_edit == "value":
        prompt_text = "Введите новое значение характеристики:"
    elif field_to_edit == "price":
        await state.set_state(AdminStates.edit_characteristic_price_input)
        prompt_text = "Введите новую цену характеристики (в копейках):"
    elif field_to_edit == "quantity":
        await state.set_state(AdminStates.edit_characteristic_quantity_input)
        prompt_text = "Введите новое количество характеристики:"
    
    # Сохраняем сообщение, которое редактируем, для возврата
    await state.update_data(message_to_edit_id=callback.message.message_id)
    await callback.message.edit_text(prompt_text)
    
    if field_to_edit != "price" and field_to_edit != "quantity": # Переход в общее состояние для текстового ввода
        await state.set_state(AdminStates.select_characteristic_to_edit)
    await callback.answer()


@router.message(AdminStates.select_characteristic_to_edit)
async def process_edit_characteristic_text_input(message: types.Message, state: FSMContext):
    data = await state.get_data()
    char_id = data.get('char_to_edit_id')
    field_to_edit = data.get('field_to_edit')
    new_value = message.text
    
    async with aiosqlite.connect("products.db") as db:
        query = f"UPDATE product_characteristics SET {field_to_edit} = ? WHERE id = ?"
        await db.execute(query, (new_value, char_id))
        await db.commit()

    await state.clear()
    await message.answer(f"Поле '{field_to_edit}' характеристики ID {char_id} обновлено.")
    # Возвращаемся к списку характеристик для этого товара
    product_id = data.get('current_product_id_char_manage')
    temp_callback = types.CallbackQuery(id='temp_id', from_user=message.from_user, chat_instance='temp_chat', data=f"manage_char_for_prod_{product_id}", message=message)
    await show_product_characteristics(temp_callback, state)


@router.message(AdminStates.edit_characteristic_price_input)
async def process_edit_characteristic_price_input(message: types.Message, state: FSMContext):
    try:
        price = int(message.text)
        if price <= 0:
            raise ValueError
        data = await state.get_data()
        char_id = data.get('char_to_edit_id')
        
        async with aiosqlite.connect("products.db") as db:
            await db.execute("UPDATE product_characteristics SET price = ? WHERE id = ?", (price, char_id))
            await db.commit()
        
        await state.clear()
        await message.answer(f"Цена характеристики ID {char_id} обновлена на {price / 100:.2f}₽.")
        product_id = data.get('current_product_id_char_manage')
        temp_callback = types.CallbackQuery(id='temp_id', from_user=message.from_user, chat_instance='temp_chat', data=f"manage_char_for_prod_{product_id}", message=message)
        await show_product_characteristics(temp_callback, state)
    except ValueError:
        await message.answer("Некорректная цена. Пожалуйста, введите целое положительное число (в копейках).")

@router.message(AdminStates.edit_characteristic_quantity_input)
async def process_edit_characteristic_quantity_input(message: types.Message, state: FSMContext):
    try:
        quantity = int(message.text)
        if quantity < 0:
            raise ValueError
        data = await state.get_data()
        char_id = data.get('char_to_edit_id')
        
        async with aiosqlite.connect("products.db") as db:
            await db.execute("UPDATE product_characteristics SET quantity = ? WHERE id = ?", (quantity, char_id))
            await db.commit()
        
        await state.clear()
        await message.answer(f"Количество характеристики ID {char_id} обновлено на {quantity} шт.")
        product_id = data.get('current_product_id_char_manage')
        temp_callback = types.CallbackQuery(id='temp_id', from_user=message.from_user, chat_instance='temp_chat', data=f"manage_char_for_prod_{product_id}", message=message)
        await show_product_characteristics(temp_callback, state)
    except ValueError:
        await message.answer("Некорректное количество. Пожалуйста, введите целое неотрицательное число.")


@router.callback_query(F.data.startswith("delete_char_"), F.from_user.id == YOUR_ADMIN_ID)
async def delete_characteristic(callback: types.CallbackQuery, state: FSMContext):
    characteristic_id = int(callback.data.split("_")[2])
    
    async with aiosqlite.connect("products.db") as db:
        # Получаем ID продукта, к которому относится характеристика, для возврата
        cursor = await db.execute("SELECT product_id FROM product_characteristics WHERE id = ?", (characteristic_id,))
        product_id = (await cursor.fetchone())[0] if await cursor.fetchone() else None

        await db.execute("DELETE FROM product_characteristics WHERE id = ?", (characteristic_id,))
        await db.commit()
    
    await callback.answer(f"Характеристика ID {characteristic_id} удалена.", show_alert=True)
    
    # Возвращаемся к списку характеристик для этого товара
    if product_id:
        temp_callback = types.CallbackQuery(id='temp_id', from_user=callback.from_user, chat_instance='temp_chat', data=f"manage_char_for_prod_{product_id}", message=callback.message)
        await show_product_characteristics(temp_callback, state)
    else:
        await callback.message.edit_text("Характеристика удалена. Вернитесь в админ-панель.")
        await state.clear()


@router.callback_query(F.data == "back_to_char_product_list", F.from_user.id == YOUR_ADMIN_ID)
async def back_to_char_product_list(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete() # Удаляем текущее сообщение
    await manage_characteristics_start(callback.message, state) # Вызываем начальную функцию управления характеристиками
    await callback.answer()


# --- Админ: Редактирование товара ---
@router.message(F.text == "✏️ Редактировать товар", F.from_user.id == YOUR_ADMIN_ID)
async def edit_product_select(message: types.Message, state: FSMContext):
    await state.clear()
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT id, name FROM products ORDER BY name ASC")
        products = await cursor.fetchall()
    
    if not products:
        await message.answer("Пока нет товаров для редактирования.")
        return

    markup_buttons = []
    for prod_id, prod_name in products:
        markup_buttons.append([InlineKeyboardButton(text=prod_name, callback_data=f"edit_prod_{prod_id}")])
    
    markup_buttons.append([InlineKeyboardButton(text="↩️ Админ-панель", callback_data="back_to_admin_menu")])
    
    await message.answer("Выберите товар для редактирования:", reply_markup=InlineKeyboardMarkup(inline_keyboard=markup_buttons))
    await state.set_state(AdminStates.edit_product_select)


@router.callback_query(F.data.startswith("edit_prod_"), AdminStates.edit_product_select, F.from_user.id == YOUR_ADMIN_ID)
async def edit_product_menu(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[2])
    await state.update_data(current_product_id=product_id)
    
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name, description, category_id, image_url FROM products WHERE id = ?", (product_id,))
        product_name, description, category_id, image_url = await cursor.fetchone()
        
        category_name = await get_category_name(category_id)

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"Название ({product_name})", callback_data="edit_product_name")],
        [InlineKeyboardButton(text=f"Описание ({description[:20]}...)", callback_data="edit_product_description")],
        [InlineKeyboardButton(text=f"Категория ({category_name})", callback_data="edit_product_category")],
        [InlineKeyboardButton(text=f"Изображение (URL)", callback_data="edit_product_image")],
        [InlineKeyboardButton(text="🗑️ Удалить товар", callback_data="delete_product_confirm")], # <-- ДОБАВЬТЕ ЭТУ СТРОКУ
        [InlineKeyboardButton(text="↩️ Назад к выбору товаров", callback_data="back_to_edit_product_list")]
    ])
    await callback.message.edit_text(f"Редактирование товара '{product_name}'. Выберите поле:", reply_markup=markup)
    await state.set_state(AdminStates.edit_product_menu)
    await callback.answer()


@router.callback_query(F.data == "delete_product_confirm", AdminStates.edit_product_menu, F.from_user.id == YOUR_ADMIN_ID)
async def delete_product_confirmation(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    product_id = data.get('current_product_id')
    
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name FROM products WHERE id = ?", (product_id,))
        product_name = (await cursor.fetchone())[0]

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Да, удалить", callback_data=f"delete_product__{product_id}")],
        [InlineKeyboardButton(text="Нет, отмена", callback_data=f"cancel_delete__{product_id}")]
    ])
    await callback.message.edit_text(f"Вы уверены, что хотите удалить товар '{product_name}' и все его характеристики?", reply_markup=markup)
    await state.set_state(AdminStates.delete_product_confirmation)
    await callback.answer()

@router.callback_query(F.data.startswith("delete_product__"), AdminStates.delete_product_confirmation, F.from_user.id == YOUR_ADMIN_ID)
async def confirm_delete_product(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("__")[1])
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name FROM products WHERE id = ?", (product_id,))
        product_name = (await cursor.fetchone())[0]
        await db.execute("DELETE FROM products WHERE id = ?", (product_id,))
        # CASCADE DELETE в product_characteristics и user_favorites должен позаботиться об остальном
        await db.commit()
    
    await state.clear()
    await callback.message.edit_text(f"Товар '{product_name}' успешно удален.")
    await callback.answer()
    await edit_product_select(callback.message, state) # Возвращаемся к списку товаров для редактирования

@router.callback_query(F.data.startswith("cancel_delete__"), AdminStates.delete_product_confirmation, F.from_user.id == YOUR_ADMIN_ID)
async def cancel_delete_product(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("__")[1])
    await state.clear()
    # Возвращаемся к меню редактирования этого товара
    temp_callback = types.CallbackQuery(id='temp_id', from_user=callback.from_user, chat_instance='temp_chat', data=f"edit_prod_{product_id}", message=callback.message)
    await edit_product_menu(temp_callback, state)
    await callback.answer("Удаление отменено.")


@router.callback_query(F.data.startswith("edit_product_"), AdminStates.edit_product_menu, F.from_user.id == YOUR_ADMIN_ID)
async def edit_product_field_prompt(callback: types.CallbackQuery, state: FSMContext):
    field_to_edit = callback.data.split("_")[2]
    await state.update_data(field_to_edit=field_to_edit)
    
    prompt_text = ""
    if field_to_edit == "name":
        prompt_text = "Введите новое название товара:"
        await state.set_state(AdminStates.edit_product_name_input)
    elif field_to_edit == "description":
        prompt_text = "Введите новое описание товара:"
        await state.set_state(AdminStates.edit_product_description_input)
    elif field_to_edit == "category":
        markup = await get_categories_markup(include_none=True, prefix="edit_product_category_")
        await callback.message.edit_text("Выберите новую категорию для товара:", reply_markup=markup)
        await state.set_state(AdminStates.edit_product_category_input)
        await callback.answer()
        return
    elif field_to_edit == "image":
        prompt_text = "Загрузите новое изображение товара или введите URL (пропустить: /skip):"
        await state.set_state(AdminStates.edit_product_image_input)
    
    await callback.message.edit_text(prompt_text)
    await callback.answer()

@router.message(AdminStates.edit_product_name_input)
async def process_edit_product_name(message: types.Message, state: FSMContext):
    data = await state.get_data()
    product_id = data.get('current_product_id')
    new_name = message.text
    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE products SET name = ? WHERE id = ?", (new_name, product_id))
        await db.commit()
    await state.clear()
    await message.answer(f"Название товара ID {product_id} обновлено на '{new_name}'.")
    # Возвращаемся в меню редактирования товара
    temp_callback = types.CallbackQuery(id='temp_id', from_user=message.from_user, chat_instance='temp_chat', data=f"edit_prod_{product_id}", message=message)
    await edit_product_menu(temp_callback, state)


@router.message(AdminStates.edit_product_description_input)
async def process_edit_product_description(message: types.Message, state: FSMContext):
    data = await state.get_data()
    product_id = data.get('current_product_id')
    new_description = message.text
    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE products SET description = ? WHERE id = ?", (new_description, product_id))
        await db.commit()
    await state.clear()
    await message.answer(f"Описание товара ID {product_id} обновлено.")
    temp_callback = types.CallbackQuery(id='temp_id', from_user=message.from_user, chat_instance='temp_chat', data=f"edit_prod_{product_id}", message=message)
    await edit_product_menu(temp_callback, state)


@router.callback_query(F.data.startswith("edit_product_category_"), AdminStates.edit_product_category_input)
async def process_edit_product_category(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    product_id = data.get('current_product_id')
    category_id = callback.data.split("_")[3]
    
    if category_id == "none":
        category_id_to_set = None
        category_name = "Без категории"
    else:
        category_id_to_set = int(category_id)
        category_name = await get_category_name(category_id_to_set)

    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE products SET category_id = ? WHERE id = ?", (category_id_to_set, product_id))
        await db.commit()
    
    await state.clear()
    await callback.message.edit_text(f"Категория товара ID {product_id} обновлена на '{category_name}'.")
    await callback.answer()
    temp_callback = types.CallbackQuery(id='temp_id', from_user=callback.from_user, chat_instance='temp_chat', data=f"edit_prod_{product_id}", message=callback.message)
    await edit_product_menu(temp_callback, state)


@router.message(AdminStates.edit_product_image_input)
async def process_edit_product_image(message: types.Message, state: FSMContext):
    data = await state.get_data()
    product_id = data.get('current_product_id')
    image_url = None
    if message.photo:
        image_url = message.photo[-1].file_id
    elif message.text and message.text.lower() != "/skip":
        image_url = message.text
    
    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE products SET image_url = ? WHERE id = ?", (image_url, product_id))
        await db.commit()
    
    await state.clear()
    await message.answer(f"Изображение товара ID {product_id} обновлено.")
    temp_callback = types.CallbackQuery(id='temp_id', from_user=message.from_user, chat_instance='temp_chat', data=f"edit_prod_{product_id}", message=message)
    await edit_product_menu(temp_callback, state)


@router.callback_query(F.data == "back_to_edit_product_list", AdminStates.edit_product_menu, F.from_user.id == YOUR_ADMIN_ID)
async def back_to_edit_product_list(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete()
    await edit_product_select(callback.message, state)
    await callback.answer()


# --- Админ: Управление категориями ---
async def get_categories_markup(include_none=False, prefix="select_category_for_product_"):
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT id, name FROM categories ORDER BY name ASC")
        categories = await cursor.fetchall()
    
    markup = InlineKeyboardMarkup(inline_keyboard=[])
    for category in categories:
        markup.inline_keyboard.append([
            InlineKeyboardButton(
                text=category[1],
                callback_data=f"select_category_for_product_{category[0]}"
            )
        ])

    if include_none:
        markup.inline_keyboard.append([
            InlineKeyboardButton(text="Без категории", callback_data="select_category_for_product_none")
        ])

    return markup



@router.message(F.text == "🗂️ Категории товаров", F.from_user.id == YOUR_ADMIN_ID)
async def manage_categories(message: types.Message, state: FSMContext):
    await state.clear()
    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить категорию", callback_data="add_category")],
        [InlineKeyboardButton(text="✏️ Переименовать категорию", callback_data="rename_category")],
        [InlineKeyboardButton(text="🗑️ Удалить категорию", callback_data="delete_category")],
        [InlineKeyboardButton(text="↩️ Админ-панель", callback_data="back_to_admin_menu")]
    ])
    await message.answer("Управление категориями товаров:", reply_markup=markup)

@router.callback_query(F.data == "add_category", F.from_user.id == YOUR_ADMIN_ID)
async def add_category_start(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(AdminStates.add_category_name)
    await callback.message.edit_text("Введите название новой категории:")
    await callback.answer()

@router.message(AdminStates.add_category_name)
async def process_add_category_name(message: types.Message, state: FSMContext):
    category_name = message.text
    async with aiosqlite.connect("products.db") as db:
        try:
            await db.execute("INSERT INTO categories (name) VALUES (?)", (category_name,))
            await db.commit()
            await message.answer(f"Категория '{category_name}' успешно добавлена!")
        except aiosqlite.IntegrityError:
            await message.answer(f"Категория '{category_name}' уже существует.")
        finally:
            await state.clear()
            await manage_categories(message, state) # Возвращаемся в меню управления категориями

@router.callback_query(F.data == "rename_category", F.from_user.id == YOUR_ADMIN_ID)
async def rename_category_select(callback: types.CallbackQuery, state: FSMContext):
    markup = await get_categories_markup(prefix="rename_cat_")
    if not markup.inline_keyboard:
        await callback.answer("Нет категорий для переименования.", show_alert=True)
        return
    await callback.message.edit_text("Выберите категорию для переименования:", reply_markup=markup)
    await state.set_state(AdminStates.rename_category_name)
    await callback.answer()

@router.callback_query(F.data.startswith("rename_cat_"), AdminStates.rename_category_name, F.from_user.id == YOUR_ADMIN_ID)
async def rename_category_prompt(callback: types.CallbackQuery, state: FSMContext):
    category_id = int(callback.data.split("_")[2])
    await state.update_data(category_id_to_rename=category_id)
    await callback.message.edit_text("Введите новое название для категории:")
    # Stay in AdminStates.rename_category_name for text input
    await callback.answer()

@router.message(AdminStates.rename_category_name)
async def process_rename_category_name(message: types.Message, state: FSMContext):
    data = await state.get_data()
    category_id = data.get('category_id_to_rename')
    new_name = message.text
    async with aiosqlite.connect("products.db") as db:
        try:
            await db.execute("UPDATE categories SET name = ? WHERE id = ?", (new_name, category_id))
            await db.commit()
            await message.answer(f"Категория успешно переименована в '{new_name}'.")
        except aiosqlite.IntegrityError:
            await message.answer(f"Категория с названием '{new_name}' уже существует.")
        finally:
            await state.clear()
            await manage_categories(message, state)

@router.callback_query(F.data == "delete_category", F.from_user.id == YOUR_ADMIN_ID)
async def delete_category_select(callback: types.CallbackQuery, state: FSMContext):
    markup = await get_categories_markup(prefix="delete_cat_")
    if not markup.inline_keyboard:
        await callback.answer("Нет категорий для удаления.", show_alert=True)
        return
    await callback.message.edit_text("Выберите категорию для удаления:", reply_markup=markup)
    await state.set_state(AdminStates.manage_categories_menu) # Переход в это состояние для ожидания выбора категории
    await callback.answer()

@router.callback_query(F.data.startswith("delete_cat_"), AdminStates.manage_categories_menu, F.from_user.id == YOUR_ADMIN_ID)
async def confirm_delete_category(callback: types.CallbackQuery, state: FSMContext):
    category_id = int(callback.data.split("_")[2])
    
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name FROM categories WHERE id = ?", (category_id,))
        category_name = (await cursor.fetchone())[0]

        # Проверяем, есть ли товары в этой категории
        cursor_products = await db.execute("SELECT COUNT(*) FROM products WHERE category_id = ?", (category_id,))
        product_count = (await cursor_products.fetchone())[0]

    if product_count > 0:
        markup = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Да, удалить и перенести товары", callback_data=f"force_delete_cat_{category_id}")],
            [InlineKeyboardButton(text="Нет, отмена", callback_data="cancel_delete_cat")]
        ])
        await callback.message.edit_text(f"Категория '{category_name}' содержит {product_count} товаров. При удалении категории, товары будут перенесены в 'Без категории'. Продолжить?", reply_markup=markup)
        await state.update_data(category_id_to_manage=category_id, category_name_to_manage=category_name)
    else:
        markup = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Да, удалить", callback_data=f"force_delete_cat_{category_id}")],
            [InlineKeyboardButton(text="Нет, отмена", callback_data="cancel_delete_cat")]
        ])
        await callback.message.edit_text(f"Вы уверены, что хотите удалить категорию '{category_name}'?", reply_markup=markup)
        await state.update_data(category_id_to_manage=category_id, category_name_to_manage=category_name)
    
    await callback.answer()


@router.callback_query(F.data == "cancel_delete_cat", F.from_user.id == YOUR_ADMIN_ID)
async def cancel_delete_category_callback(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("Удаление категории отменено.")
    await callback.answer()
    await manage_categories(callback.message, state)


@router.callback_query(F.data.startswith("force_delete_cat_"), F.from_user.id == YOUR_ADMIN_ID)
async def execute_delete_category(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    category_id = data.get('category_id_to_manage')
    category_name = data.get('category_name_to_manage')

    async with aiosqlite.connect("products.db") as db:
        # 1. Все товары из этой категории делаем "Без категории"
        await db.execute("UPDATE products SET category_id = NULL WHERE category_id = ?", (category_id,))
        # 2. Удаляем саму категорию
        await db.execute("DELETE FROM categories WHERE id = ?", (category_id,))
        await db.commit()

    await state.clear()
    await callback.message.edit_text(f"Категория '{category_name}' успешно удалена.")
    
    markup = await get_categories_markup()
    await callback.message.answer("Меню категорий обновлено:", reply_markup=markup)
    await callback.answer()


# --- Вспомогательные функции для работы с корзиной в БД ---
async def load_user_cart_from_db(user_id):
    """Загружает корзину пользователя из базы данных."""
    try:
        async with aiosqlite.connect("products.db") as db:
            cursor = await db.execute("SELECT cart_items FROM user_carts WHERE user_id = ?", (user_id,))
            result = await cursor.fetchone()
            if result and result[0]:
                cart_data = json.loads(result[0])
                print(f"DEBUG: Loaded cart for user {user_id}: {cart_data}")
                return cart_data
            print(f"DEBUG: No cart found or empty cart for user {user_id}")
            return []
    except Exception as e:
        print(f"ERROR in load_user_cart_from_db: {str(e)}")
        return []

async def save_user_cart_to_db(user_id, cart):
    """Сохраняет корзину пользователя в базу данных."""
    try:
        async with aiosqlite.connect("products.db") as db:
            cart_json = json.dumps(cart)
            current_time = datetime.now().strftime('%d-%m-%Y')
            await db.execute(
                "INSERT OR REPLACE INTO user_carts (user_id, cart_items, last_updated) VALUES (?, ?, ?)",
                (user_id, cart_json, current_time)
            )
            await db.commit()
            print(f"DEBUG: Saved cart for user {user_id}: {cart_json} at {current_time}")
    except Exception as e:
        print(f"ERROR in save_user_cart_to_db: {str(e)}")
        raise

async def init_db():
    async with aiosqlite.connect("products.db") as db:
        # Таблица категорий
        await db.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE
            )""")
        
        # Обновленная таблица products (без price и quantity)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                description TEXT,
                category_id INTEGER,
                image_url TEXT,
                FOREIGN KEY (category_id) REFERENCES categories (id)
            )""")

        # НОВАЯ ТАБЛИЦА: product_characteristics
        await db.execute("""
            CREATE TABLE IF NOT EXISTS product_characteristics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                name TEXT NOT NULL,         -- Например, "Объем", "Цвет"
                value TEXT NOT NULL,        -- Например, "5мл", "Красный"
                price INTEGER NOT NULL,     -- Цена для этой характеристики
                quantity INTEGER NOT NULL,  -- Количество для этой характеристики
                FOREIGN KEY (product_id) REFERENCES products (id) ON DELETE CASCADE
            )""")
        
        
        # Обновленная таблица order_items (ссылается на characteristic_id)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                characteristic_id INTEGER, -- Теперь ссылаемся на характеристику
                quantity INTEGER,
                FOREIGN KEY (characteristic_id) REFERENCES product_characteristics (id) ON DELETE CASCADE
            )""")
        
        # Таблица для избранных товаров (пока оставим на уровне product_id для простоты)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS user_favorites (
                user_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, product_id),
                FOREIGN KEY (product_id) REFERENCES products (id) ON DELETE CASCADE
            )""")

        # МИГРАЦИЯ СХЕМЫ: Удаление старых колонок 'price' и 'quantity' из 'products'
        cursor = await db.execute("PRAGMA table_info(products)")
        columns = [col[1] for col in await cursor.fetchall()]
        
        if 'price' in columns:
            print("DEBUG: 'price' column found in 'products'. Attempting to drop it.")
            # SQLite не поддерживает прямое DROP COLUMN, нужна миграция
            await db.execute("ALTER TABLE products RENAME TO _old_products_with_price")
            await db.execute("""
                CREATE TABLE products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT,
                    description TEXT,
                    category_id INTEGER,
                    image_url TEXT
                )""")
            await db.execute("""
                INSERT INTO products (id, name, description, category_id, image_url)
                SELECT id, name, description, category_id, image_url FROM _old_products_with_price
            """)
            await db.execute("DROP TABLE _old_products_with_price")
            print("DEBUG: 'price' column removed and table recreated.")

        if 'quantity' in columns:
            print("DEBUG: 'quantity' column found in 'products'. Attempting to drop it.")
            # Повторная миграция, если 'price' уже удаляли, или первая, если только 'quantity'
            cursor_new = await db.execute("PRAGMA table_info(products)")
            current_columns = [col[1] for col in await cursor_new.fetchall()]
            if 'quantity' in current_columns: # Проверяем еще раз, вдруг уже удалилась с price
                await db.execute("ALTER TABLE products RENAME TO _old_products_with_quantity")
                await db.execute("""
                    CREATE TABLE products (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT,
                        description TEXT,
                        category_id INTEGER,
                        image_url TEXT
                    )""")
                await db.execute("""
                    INSERT INTO products (id, name, description, category_id, image_url)
                    SELECT id, name, description, category_id, image_url FROM _old_products_with_quantity
                """)
                await db.execute("DROP TABLE _old_products_with_quantity")
                print("DEBUG: 'quantity' column removed and table recreated.")

        # Проверка и добавление колонки category_id в products, если её нет (старая проверка, можно оставить)
        cursor_final = await db.execute("PRAGMA table_info(products)")
        final_columns = [col[1] for col in await cursor_final.fetchall()]
        if 'category_id' not in final_columns:
            print("DEBUG: 'category_id' column missing in 'products'. Adding it.")
            await db.execute("ALTER TABLE products ADD COLUMN category_id INTEGER")
            print("DEBUG: 'category_id' column added to 'products'.")
        
        # Проверка и добавление колонки image_url в products, если её нет (старая проверка, можно оставить)
        if 'image_url' not in final_columns:
            print("DEBUG: 'image_url' column missing in 'products'. Adding it.")
            await db.execute("ALTER TABLE products ADD COLUMN image_url TEXT")
            print("DEBUG: 'image_url' column added to 'products'.")
        
        # Проверка наличия таблицы user_carts и колонки last_updated (оставить как есть)
        cursor = await db.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_carts'")
        user_carts_exists = await cursor.fetchone()
        if not user_carts_exists:
            print("DEBUG: Creating new 'user_carts' table with 'last_updated' column.")
            await db.execute("""
                CREATE TABLE user_carts (
                    user_id INTEGER PRIMARY KEY,
                    cart_items TEXT,
                    last_updated TEXT DEFAULT CURRENT_TIMESTAMP
                )""")
        else:
            cursor = await db.execute("PRAGMA table_info(user_carts)")
            columns = [col[1] for col in await cursor.fetchall()]
            if 'last_updated' not in columns:
                print("DEBUG: 'last_updated' column missing in 'user_carts'. Performing schema migration.")
                await db.execute("ALTER TABLE user_carts RENAME TO _old_user_carts")
                await db.execute("""
                    CREATE TABLE user_carts (
                        user_id INTEGER PRIMARY KEY,
                        cart_items TEXT,
                        last_updated TEXT DEFAULT CURRENT_TIMESTAMP
                    )""")
                await db.execute("""
                    INSERT INTO user_carts (user_id, cart_items, last_updated)
                    SELECT user_id, cart_items, CURRENT_TIMESTAMP FROM _old_user_carts
                """)
                await db.execute("DROP TABLE _old_user_carts")
                print("DEBUG: Schema migration for 'user_carts' complete.")

        # Добавляем стандартные категории, если их нет (оставить как есть)
        initial_categories = ["Женский парфюм", "Мужской парфюм", "Унисекс парфюм"]
        for category_name in initial_categories:
            await db.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (category_name,))
        
        await db.commit()

async def get_category_name(category_id):
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name FROM categories WHERE id = ?", (category_id,))
        result = await cursor.fetchone()
        return result[0] if result else "Без категории"

# <--- НОВЫЙ БЛОК: Функции для работы с избранным
async def is_product_in_favorites(user_id: int, product_id: int) -> bool:
    """Проверяет, находится ли товар в избранном у пользователя."""
    try:
        async with aiosqlite.connect("products.db") as db:
            cursor = await db.execute(
                "SELECT 1 FROM user_favorites WHERE user_id = ? AND product_id = ?",
                (user_id, product_id)
            )
            return await cursor.fetchone() is not None
    except Exception as e:
        print(f"DEBUG: Ошибка в is_product_in_favorites: {e}")
        return False

# --- Логика каталога (полностью переработана) ---

@router.message(F.text == "📋 Каталог")
async def show_categories(message: types.Message, state: FSMContext):
    await state.clear()
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("""
            SELECT c.id, c.name, (SELECT COUNT(*) FROM products p WHERE p.category_id = c.id) as product_count
            FROM categories c
        """)
        categories_with_counts = await cursor.fetchall()
    
    buttons = []
    if categories_with_counts:
        for cat_id, cat_name, product_count in categories_with_counts:
            buttons.append([InlineKeyboardButton(text=f"{cat_name} ({product_count})", callback_data=f"show_cat_{cat_id}_page_0")])
    
    # <--- ИЗМЕНЕНО: Добавлена кнопка "Избранное"
    buttons.append([InlineKeyboardButton(text="❤️ Избранное", callback_data="show_fav_page_0")])

    await message.answer("Выберите категорию товаров или просмотрите избранное:", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))

# --- Генераторы клавиатур для карточек
def get_product_card_markup(product_id: int, price: int, is_favorite: bool) -> InlineKeyboardMarkup:
    """Создает стандартную клавиатуру для карточки товара."""
    favorite_text = "💔 Удалить из избранного" if is_favorite else "❤️ Добавить в избранное"
    favorite_callback = f"fav_rem_{product_id}" if is_favorite else f"fav_add_{product_id}"
    buttons = [
        [InlineKeyboardButton(text=f"Купить ({price / 100:.2f}₽)", callback_data=f"buy_prod_{product_id}")],
        [InlineKeyboardButton(text=favorite_text, callback_data=favorite_callback)]
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def get_quantity_selector_markup(characteristic_id: int, current_qty: int, max_qty: int) -> InlineKeyboardMarkup:
    """Создает клавиатуру для выбора количества для выбранной характеристики."""
    buttons = [
        [
            InlineKeyboardButton(text="-", callback_data=f"qty-char-change_dec_{characteristic_id}_{current_qty}"),
            InlineKeyboardButton(text=f"{current_qty} шт.", callback_data="noop"),
            InlineKeyboardButton(text="+", callback_data=f"qty-char-change_inc_{characteristic_id}_{current_qty}")
        ],
        [InlineKeyboardButton(text=f"✅ Добавить в корзину", callback_data=f"cart_add_char_{characteristic_id}_{current_qty}")],
        # Изменён формат callback данных для кнопки "Назад"
        [InlineKeyboardButton(text="⬅️ Назад к выбору характеристики", callback_data=f"cancel_qty_selector_char_{characteristic_id}")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)

# --- Генераторы клавиатур для карточек (Обновленная) ---
def get_product_card_markup(product_id: int, has_characteristics: bool, is_favorite: bool) -> InlineKeyboardMarkup:
    """
    Создает стандартную клавиатуру для карточки товара.
    has_characteristics: True, если у товара есть характеристики.
    """
    favorite_text = "💔 Удалить из избранного" if is_favorite else "❤️ Добавить в избранное"
    favorite_callback = f"fav_rem_{product_id}" if is_favorite else f"fav_add_{product_id}"

    buttons = []
    if has_characteristics:
        buttons.append([InlineKeyboardButton(text="Выбрать параметры", callback_data=f"select_char_{product_id}")])
    else:
        # Если нет характеристик, возможно, товар должен быть "недоступен" или иметь стандартную цену/кол-во
        # Для простоты, здесь мы не добавляем кнопку "Купить", если нет характеристик
        pass
        
    buttons.append([InlineKeyboardButton(text=favorite_text, callback_data=favorite_callback)])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def get_characteristic_selection_markup(product_id: int, characteristics: list) -> InlineKeyboardMarkup:
    """Создает клавиатуру для выбора характеристики товара."""
    buttons = []
    for char_id, char_name, char_value, char_price, char_qty in characteristics:
        if char_qty > 0: # Показываем только те характеристики, которые есть в наличии
            buttons.append([InlineKeyboardButton(text=f"{char_name}: {char_value} ({char_price / 100:.2f}₽)", callback_data=f"buy_char_{char_id}")])
        else:
            buttons.append([InlineKeyboardButton(text=f"{char_name}: {char_value} (Нет в наличии)", callback_data="noop_char_no_stock")])

    buttons.append([InlineKeyboardButton(text="⬅️ Назад к товару", callback_data=f"back_to_product_view_{product_id}")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def get_quantity_selector_markup(characteristic_id: int, current_qty: int, max_qty: int) -> InlineKeyboardMarkup:
    """Создает клавиатуру для выбора количества для выбранной характеристики."""
    buttons = [
        [
            InlineKeyboardButton(text="-", callback_data=f"qty-char-change_dec_{characteristic_id}_{current_qty}"),
            InlineKeyboardButton(text=f"{current_qty} шт.", callback_data="noop"),
            InlineKeyboardButton(text="+", callback_data=f"qty-char-change_inc_{characteristic_id}_{current_qty}")
        ],
        [InlineKeyboardButton(text=f"✅ Добавить в корзину", callback_data=f"cart_add_char_{characteristic_id}_{current_qty}")],
        [InlineKeyboardButton(text="⬅️ Назад к выбору характеристики", callback_data=f"cancel_qty_selector_char_{characteristic_id}")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


async def send_products_page(message: types.Message, page: int, state: FSMContext, category_id: int = None, from_favorites: bool = False):
    offset = page * PRODUCTS_PER_PAGE
    user_id = message.chat.id
    
    async with aiosqlite.connect("products.db") as db:
        if from_favorites:
            cursor_count = await db.execute("SELECT COUNT(*) FROM user_favorites WHERE user_id = ?", (user_id,))
            total_products = (await cursor_count.fetchone())[0]
            cursor = await db.execute(
                """SELECT p.id, p.name, p.description, p.image_url
                   FROM products p JOIN user_favorites uf ON p.id = uf.product_id
                   WHERE uf.user_id = ? LIMIT ? OFFSET ?""",
                (user_id, PRODUCTS_PER_PAGE, offset)
            )
            products = await cursor.fetchall()
            page_title = f"❤️ Избранное (Страница {page + 1})"
            base_callback = "show_fav"
        else:
            cursor_count = await db.execute("SELECT COUNT(*) FROM products WHERE category_id = ?", (category_id,))
            total_products = (await cursor_count.fetchone())[0]
            cursor = await db.execute(
                "SELECT id, name, description, image_url FROM products WHERE category_id = ? LIMIT ? OFFSET ?",
                (category_id, PRODUCTS_PER_PAGE, offset)
            )
            products = await cursor.fetchall()
            category_name = await get_category_name(category_id)
            page_title = f"📋 Каталог - {category_name} (Страница {page + 1})"
            base_callback = f"show_cat_{category_id}"

    if not products:
        empty_message = "Ваше избранное пока пусто 💔. Добавьте сюда товары, которые вам понравились, чтобы не потерять их и легко вернуться к покупкам! Загляните в наш 📋 Каталог — там точно найдётся что-то особенное для вас!" if from_favorites else "В этой категории пока нет товаров."
        await message.answer(empty_message, 
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="📋 Каталог", callback_data="show_main_categories")]]))
        return

    await message.answer(f"**{page_title}**", parse_mode="Markdown")

    for prod_id, name, desc, image_url in products: # Изменено: нет price, qty
        caption = f"**{name}**\n\n{desc or 'Описание отсутствует'}"
        
        # Проверяем, есть ли у товара характеристики
        async with aiosqlite.connect("products.db") as db_inner:
            cursor_char = await db_inner.execute("SELECT COUNT(*) FROM product_characteristics WHERE product_id = ?", (prod_id,))
            has_characteristics = (await cursor_char.fetchone())[0] > 0
        
        is_fav = await is_product_in_favorites(user_id, prod_id)
        markup = get_product_card_markup(prod_id, has_characteristics, is_fav)

        if not has_characteristics: # Если нет характеристик, считаем товар недоступным для покупки
            caption += "\n\n**Нет доступных характеристик**"
            markup = None # Убираем кнопки, если нельзя купить
            
        if image_url:
            try:
                await bot.send_photo(chat_id=user_id, photo=image_url, caption=caption, reply_markup=markup, parse_mode="Markdown")
            except Exception as e:
                await bot.send_message(user_id, caption + "\n\n(Не удалось загрузить изображение)", reply_markup=markup, parse_mode="Markdown")
        else:
            await bot.send_message(user_id, caption + "\n\n(Изображение отсутствует)", reply_markup=markup, parse_mode="Markdown")

    nav_buttons = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton(text="⬅️ Назад", callback_data=f"{base_callback}_page_{page - 1}"))
    if (page + 1) * PRODUCTS_PER_PAGE < total_products:
        nav_buttons.append(InlineKeyboardButton(text="Далее ➡️", callback_data=f"{base_callback}_page_{page + 1}"))

    if nav_buttons:
        pagination_markup = InlineKeyboardMarkup(inline_keyboard=[nav_buttons])
        await message.answer("Перейти на другую страницу:", reply_markup=pagination_markup)
 
@router.callback_query(F.data.startswith("show_cat_"))
async def navigate_categories(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.delete()
    parts = callback.data.split("_")
    category_id = int(parts[2])
    page = int(parts[4])
    await send_products_page(callback.message, page, state, category_id=category_id)
    await callback.answer()

@router.callback_query(F.data.startswith("show_fav_"))
async def navigate_favorites(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.delete()
    page = int(callback.data.split("_")[3])
    await send_products_page(callback.message, page, state, from_favorites=True)
    await callback.answer()

@router.callback_query(F.data.startswith("select_char_"))
async def show_characteristic_selector(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[2])
    await state.update_data(current_product_id=product_id) # Сохраняем ID товара для возврата

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT id, name, value, price, quantity FROM product_characteristics WHERE product_id = ? ORDER BY price ASC", (product_id,))
        characteristics = await cursor.fetchall()
        
        cursor_product = await db.execute("SELECT name, description, image_url FROM products WHERE id = ?", (product_id,))
        product_name, product_description, product_image_url = await cursor_product.fetchone()

    if not characteristics:
        await callback.answer("У этого товара нет доступных характеристик.", show_alert=True)
        # Вернуться на предыдущую страницу или обновить карточку товара
        is_fav = await is_product_in_favorites(callback.from_user.id, product_id)
        markup = get_product_card_markup(product_id, False, is_fav) # False, так как нет характеристик
        await callback.message.edit_reply_markup(reply_markup=markup)
        return
    
    # Редактируем сообщение с карточкой товара, чтобы показать выбор характеристик
    caption = f"**{product_name}**\n\n{product_description or 'Описание отсутствует'}\n\nВыберите вариант:"
    
    markup = get_characteristic_selection_markup(product_id, characteristics)
    
    # Если это фото, редактируем подпись и кнопки
    if callback.message.photo:
        await callback.message.edit_caption(caption=caption, reply_markup=markup, parse_mode="Markdown")
    else: # Если это текстовое сообщение
        await callback.message.edit_text(text=caption, reply_markup=markup, parse_mode="Markdown")
        
    await callback.answer()

@router.callback_query(F.data.startswith("buy_char_"))
async def show_quantity_selector_for_characteristic(callback: types.CallbackQuery, state: FSMContext):
    characteristic_id = int(callback.data.split("_")[2])
    await state.update_data(current_characteristic_id=characteristic_id) # Сохраняем ID характеристики

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name, value, price, quantity FROM product_characteristics WHERE id = ?", (characteristic_id,))
        char_name, char_value, char_price, char_qty = await cursor.fetchone()
        
        
        # Получаем ID продукта, к которому относится характеристика
        cursor_product_id = await db.execute("SELECT product_id FROM product_characteristics WHERE id = ?", (characteristic_id,))
        product_id = (await cursor_product_id.fetchone())[0]
        await state.update_data(current_product_id=product_id) # Обновляем product_id в состоянии

    if not char_qty or char_qty == 0:
        await callback.answer("Этого варианта нет в наличии!", show_alert=True)
        return
    
    # --- ИСПРАВЛЕННЫЙ КОД НАЧИНАЕТСЯ ЗДЕСЬ ---
    # 1. Инициализируем базовое содержимое подписи/текста пустым значением
    base_caption_content = ""
    
    # 2. Безопасно получаем базовый текст из callback.message
    # Проверяем, есть ли у сообщения подпись (для фото/видео)
    if callback.message.caption:
        # Разбиваем подпись по "Выберите вариант:" и берем часть до него
        parts = callback.message.caption.split('Выберите вариант:', 1)
        base_caption_content = parts[0].strip() # .strip() убирает лишние пробелы
    elif callback.message.text:
        # Если это обычное текстовое сообщение
        # Разбиваем текст по "Выберите вариант:" и берем часть до него
        parts = callback.message.text.split('Выберите вариант:', 1)
        base_caption_content = parts[0].strip()

    # 3. Формируем полную подпись сообщения.
    # Эта часть теперь гарантированно выполняется, если функция не вернулась выше.
    caption = (
        f"{base_caption_content}\n\n"
        f"Выбран: **{char_name}: {char_value}**\n"
        f"Цена: {char_price / 100:.2f}₽\n"
        f"В наличии: {char_qty} шт.\n\n"
        f"Выберите количество:"
    )
    # --- ИСПРАВЛЕННЫЙ КОД ЗАВЕРШАЕТСЯ ЗДЕСЬ ---
    
    markup = get_quantity_selector_markup(characteristic_id, 1, char_qty)

    # Используем edit_caption для фото и edit_text для обычных сообщений
    if callback.message.photo:
        await callback.message.edit_caption(caption=caption, reply_markup=markup, parse_mode="Markdown")
    else:
        await callback.message.edit_text(text=caption, reply_markup=markup, parse_mode="Markdown")
        
    await callback.answer()

@router.callback_query(F.data.startswith("noop_char_no_stock"))
async def noop_char_no_stock_callback(callback: types.CallbackQuery):
    await callback.answer("Этого варианта нет в наличии.", show_alert=True)

@router.callback_query(F.data.startswith("back_to_product_view_"))
async def back_to_product_view_callback(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[4])
    await state.update_data(current_product_id=product_id) # Убедимся, что product_id в состоянии
    
    async with aiosqlite.connect("products.db") as db:
        cursor_product = await db.execute("SELECT name, description, image_url FROM products WHERE id = ?", (product_id,))
        product_name, product_description, product_image_url = await cursor_product.fetchone()
        
        cursor_char_count = await db.execute("SELECT COUNT(*) FROM product_characteristics WHERE product_id = ?", (product_id,))
        has_characteristics = (await cursor_char_count.fetchone())[0] > 0

    caption = f"**{product_name}**\n\n{product_description or 'Описание отсутствует'}"
    is_fav = await is_product_in_favorites(callback.from_user.id, product_id)
    markup = get_product_card_markup(product_id, has_characteristics, is_fav)

    if not has_characteristics:
        caption += "\n\n**Нет доступных характеристик**"
        markup = None

    if callback.message.photo:
        await callback.message.edit_caption(caption=caption, reply_markup=markup, parse_mode="Markdown")
    else:
        await callback.message.edit_text(caption=caption, reply_markup=markup, parse_mode="Markdown")
    
    await callback.answer()


@router.callback_query(F.data.startswith("cancel_qty_selector_char_"))
async def cancel_quantity_selector_char(callback: types.CallbackQuery, state: FSMContext):
    characteristic_id = int(callback.data.split("_")[3])
    
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT product_id FROM product_characteristics WHERE id = ?", (characteristic_id,))
        product_id = (await cursor.fetchone())[0]

        cursor_chars = await db.execute("SELECT id, name, value, price, quantity FROM product_characteristics WHERE product_id = ? ORDER BY price ASC", (product_id,))
        characteristics = await cursor_chars.fetchall()

        cursor_product = await db.execute("SELECT name, description, image_url FROM products WHERE id = ?", (product_id,))
        product_name, product_description, product_image_url = await cursor_product.fetchone()

    caption = f"**{product_name}**\n\n{product_description or 'Описание отсутствует'}\n\nВыберите вариант:"
    markup = get_characteristic_selection_markup(product_id, characteristics)
    
    if callback.message.photo:
        await callback.message.edit_caption(caption=caption, reply_markup=markup, parse_mode="Markdown")
    else:
        await callback.message.edit_text(caption=caption, reply_markup=markup, parse_mode="Markdown")
    
    await callback.answer()

@router.callback_query(F.data.startswith("qty-char-change_")) # ИЗМЕНЕНО: Фильтр для нового формата характеристик
async def change_quantity_for_characteristic(callback: types.CallbackQuery, state: FSMContext):
    try:
        _, _, action, characteristic_id_str, qty_str = callback.data.split("_")
    except ValueError:
        await callback.answer("Ошибка! Попробуйте снова.", show_alert=True)
        return

    characteristic_id = int(characteristic_id_str)
    current_qty = int(qty_str)

    new_qty = current_qty + 1 if action == "inc" else current_qty - 1

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT quantity FROM product_characteristics WHERE id = ?", (characteristic_id,))
        result = await cursor.fetchone()

    if not result:
        await callback.answer("Извините, этот вариант товара больше недоступен.", show_alert=True)
        # Обновить сообщение, чтобы убрать кнопки покупки
        if callback.message.photo:
            await callback.message.edit_caption(
                caption=f"{callback.message.caption.split('Выберите количество:')[0]}\n\n**ВАРИАНТ БОЛЬШЕ НЕДОСТУПЕН**",
                reply_markup=None
            )
        else:
            await callback.message.edit_text(
                text=f"{callback.message.text.split('Выберите количество:')[0]}\n\n**ВАРИАНТ БОЛЬШЕ НЕДОСТУПЕН**",
                reply_markup=None
            )
        return

    max_qty = result[0]

    if new_qty < 1:
        new_qty = 1
    if new_qty > max_qty:
        await callback.answer(f"В наличии только {max_qty} шт.", show_alert=True)
        new_qty = max_qty

    if new_qty != current_qty:
        markup = get_quantity_selector_markup(characteristic_id, new_qty, max_qty)
        await callback.message.edit_reply_markup(reply_markup=markup)

    await callback.answer() 
 
@router.callback_query(F.data.startswith("qty-change_"))
async def change_quantity(callback: types.CallbackQuery, state: FSMContext): # <--- ИЗМЕНЕНИЕ ЗДЕСЬ: добавлен 'state: FSMContext'
    parts = callback.data.split("_")
    action = parts[1] # 'plus' или 'minus'
    characteristic_id = int(parts[2]) # Теперь это ID характеристики
    current_qty_in_selector = int(parts[3])

    async with aiosqlite.connect("products.db") as db:
        # Получаем максимальное количество для данной характеристики
        cursor = await db.execute("SELECT quantity FROM product_characteristics WHERE id = ?", (characteristic_id,))
        max_qty = (await cursor.fetchone())[0]

    new_qty = current_qty_in_selector
    if action == "plus":
        if current_qty_in_selector < max_qty:
            new_qty += 1
    elif action == "minus":
        if current_qty_in_selector > 1:
            new_qty -= 1
    
    # Получаем product_id из состояния (теперь 'state' определен)
    state_data = await state.get_data() 
    # Предполагается, что current_product_id или similar_id был сохранен в состоянии
    # Возможно, тебе здесь нужен не product_id, а characteristic_id для дальнейших операций.
    # Если product_id не используется ниже, эту строку можно и удалить, но лучше оставить, если он понадобится.
    # product_id_from_state = state_data.get('current_product_id') 

    # Проверяем, изменилось ли количество
    if new_qty != current_qty_in_selector:
        # Вызываем обновленную функцию-генератор
        # get_quantity_selector_markup теперь принимает characteristic_id, а не product_id
        markup = get_quantity_selector_markup(characteristic_id, new_qty, max_qty) 
        await callback.message.edit_reply_markup(reply_markup=markup)
    await callback.answer() 
 
@router.callback_query(F.data.startswith("cart_add_char_"))
async def add_characteristic_to_cart(callback: types.CallbackQuery, state: FSMContext):
    _, _, characteristic_id_str, qty_str = callback.data.split("_")
    characteristic_id = int(characteristic_id_str)
    quantity_to_add = int(qty_str)
    user_id = callback.from_user.id

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT quantity, name, value, price FROM product_characteristics WHERE id = ?", (characteristic_id,))
        result = await cursor.fetchone()
        if not result:
            await callback.answer("Этот вариант товара больше недоступен.", show_alert=True)
            return

        available_qty, char_name, char_value, char_price = result

    cart = await load_user_cart_from_db(user_id)
    
    current_qty_in_cart = 0
    item_index = -1
    # Корзина теперь хранит (characteristic_id, quantity)
    for i, (cid, qty) in enumerate(cart):
        if cid == characteristic_id:
            current_qty_in_cart = qty
            item_index = i
            break
            
    if current_qty_in_cart + quantity_to_add > available_qty:
        await callback.answer(f"Недостаточно товара. Доступно: {available_qty} шт.", show_alert=True)
        return

    if item_index != -1:
        cart[item_index] = (characteristic_id, current_qty_in_cart + quantity_to_add)
    else:
        cart.append((characteristic_id, quantity_to_add))
        
    await save_user_cart_to_db(user_id, cart)
    await callback.answer(f"Добавлено в корзину: {quantity_to_add} шт. ({char_name}: {char_value})", show_alert=True)
    
    # Возвращаемся к выбору количества для этой же характеристики, чтобы можно было добавить еще
    # Или можно вернуться к списку характеристик для этого товара
    # Для простоты, оставим на экране выбора количества, но обновим его
    markup = get_quantity_selector_markup(characteristic_id, quantity_to_add + current_qty_in_cart, available_qty)
    # Обновим текст сообщения, если пользователь добавил в корзину несколько раз
    updated_caption = f"{callback.message.caption.split('В наличии:')[0]}В наличии: {available_qty} шт.\n\nКоличество в корзине: {quantity_to_add + current_qty_in_cart} шт.\nВыберите количество:"
    
    if callback.message.photo:
        await callback.message.edit_caption(caption=updated_caption, reply_markup=markup, parse_mode="Markdown")
    else:
        await callback.message.edit_text(text=updated_caption, reply_markup=markup, parse_mode="Markdown")
        

@router.callback_query(F.data.startswith(("fav_add_", "fav_rem_")))
async def handle_favorite_toggle(callback: types.CallbackQuery):
    try:
        # Получаем product_id из callback data
        product_id = int(callback.data.split("_")[2])
        user_id = callback.from_user.id
        
        async with aiosqlite.connect("products.db") as db:
            is_favorite = await is_product_in_favorites(user_id, product_id)
            
            if is_favorite:
                # Удаляем из избранного
                await db.execute(
                    "DELETE FROM user_favorites WHERE user_id = ? AND product_id = ?",
                    (user_id, product_id)
                )
                await db.commit()
                message = "Товар удален из избранного"
                is_now_favorite = False
            else:
                # Проверяем существование товара
                cursor = await db.execute("SELECT 1 FROM products WHERE id = ?", (product_id,))
                if await cursor.fetchone():
                    # Добавляем в избранное
                    await db.execute(
                        "INSERT OR IGNORE INTO user_favorites (user_id, product_id) VALUES (?, ?)",
                        (user_id, product_id)
                    )
                    await db.commit()
                    message = "Товар добавлен в избранное"
                    is_now_favorite = True
                else:
                    await callback.answer("Товар не найден")
                    return

        # Обновляем кнопки в сообщении
        markup = get_product_card_markup(product_id, True, is_now_favorite)  # Предполагаем, что товар имеет характеристики
        
        if callback.message.photo:
            await callback.message.edit_reply_markup(reply_markup=markup)
        else:
            await callback.message.edit_reply_markup(reply_markup=markup)
            
        await callback.answer(message, show_alert=True)
        
    except Exception as e:
        print(f"ERROR in handle_favorite_toggle: {str(e)}")
        await callback.answer("Произошла ошибка при обработке избранного", show_alert=True)


@router.callback_query(F.data == "noop")
async def noop_callback(callback: types.CallbackQuery):
    await callback.answer()

@dp.callback_query(F.data == "show_main_categories")
async def navigate_to_main_categories(callback: types.CallbackQuery, state: FSMContext):
    # Эта функция вызывает show_categories, которая отображает список категорий.
    await show_categories(callback.message, state)
    await callback.message.delete() 
    await callback.answer()


# --- Логика корзины и оформления заказа ---
@router.message(F.text == "🛒 Корзина")
async def view_cart(message: types.Message):
    user_id = message.from_user.id
    print(f"DEBUG: view_cart called for user {user_id}")
    cart = await load_user_cart_from_db(user_id)
    # user_cart[user_id] = cart # Эта глобальная переменная больше не используется как основной источник корзины

    if not cart:
        await message.answer("Ваша корзина пуста")
        return
    
    text = "🛒 **Ваша корзина**:\n\n"
    total_price = 0
    buttons = []
    async with aiosqlite.connect("products.db") as db:
        for char_id, qty in cart: # Теперь получаем characteristic_id
            cursor = await db.execute(
                """SELECT pc.name, pc.value, pc.price, p.name as product_name
                   FROM product_characteristics pc
                   JOIN products p ON pc.product_id = p.id
                   WHERE pc.id = ?""", (char_id,)
            )
            result = await cursor.fetchone()
            if result:
                char_name, char_value, char_price, product_name = result
                item_total = char_price * qty / 100
                display_name = f"{product_name} ({char_name}: {char_value})"
                text += f"**{display_name}** — {qty} шт — {item_total:.2f}₽\n"
                total_price += item_total
                buttons.append([
                    InlineKeyboardButton(text=f"✏️ Изменить кол-во ({display_name})", callback_data=f"change_qty_char_item_id_{char_id}"),
                    InlineKeyboardButton(text=f"🗑️ Удалить ({display_name})", callback_data=f"remove_char_item_id_{char_id}")
                ])
            else:
                text += f"Неизвестный товар ID {char_id} (удалён из каталога) — {qty} шт\n"
                buttons.append([
                    InlineKeyboardButton(text=f"🗑️ Удалить (Неизвестный товар ID {char_id})", callback_data=f"remove_char_item_id_{char_id}")
                ])
    
    text += f"\n**Итого**: {total_price:.2f}₽"
    
    markup = ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📋 Каталог")],
        [KeyboardButton(text="✅ Оформить заказ")],
        [KeyboardButton(text="🗑️ Очистить корзину")],
        [KeyboardButton(text="↩️ Главное меню")]
    ], resize_keyboard=True)
    
    await message.answer(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode="Markdown")
    await message.answer("Выберите действие:", reply_markup=markup)

@router.message(F.text == "🗑️ Очистить корзину")
async def clear_cart(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    await save_user_cart_to_db(user_id, [])
    user_cart[user_id] = []
    await message.answer("Ваша корзина очищена.")
    await state.clear()
    await start(message)

@router.callback_query(F.data.startswith("remove_char_item_id_")) # ИЗМЕНЕНО: Новое имя callback
async def cart_remove_item_confirm(callback: types.CallbackQuery, state: FSMContext):
    characteristic_id_to_remove = int(callback.data.split("_")[4]) # ИЗМЕНЕНО: Индекс для characteristic_id
    user_id = callback.from_user.id
    
    cart = await load_user_cart_from_db(user_id)
    # Удаляем элемент по characteristic_id
    cart = [(cid, qty) for cid, qty in cart if cid != characteristic_id_to_remove]
    await save_user_cart_to_db(user_id, cart)
    
    await callback.message.edit_text("Товар успешно удалён из корзины.", reply_markup=None)
    await view_cart(callback.message)
    await callback.answer()
    await state.clear()

@router.callback_query(F.data.startswith("change_qty_char_item_id_")) # ИЗМЕНЕНО: Новое имя callback
async def cart_change_qty_selected(callback: types.CallbackQuery, state: FSMContext):
    characteristic_id_to_change = int(callback.data.split("_")[5]) # ИЗМЕНЕНО: Индекс для characteristic_id
    await state.update_data(characteristic_id_to_change_qty=characteristic_id_to_change) # ИЗМЕНЕНО: Сохраняем characteristic_id
    
    user_id = callback.from_user.id
    cart = await load_user_cart_from_db(user_id)
    
    current_qty = 0
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute(
            """SELECT pc.name, pc.value, p.name as product_name 
               FROM product_characteristics pc 
               JOIN products p ON pc.product_id = p.id
               WHERE pc.id = ?""", (characteristic_id_to_change,)
        )
        result = await cursor.fetchone()
        if result:
            char_name, char_value, product_name = result
            display_name = f"{product_name} ({char_name}: {char_value})"
        else:
            display_name = f"Неизвестный товар ID {characteristic_id_to_change}"
    
    for cid, qty in cart:
        if cid == characteristic_id_to_change:
            current_qty = qty
            break

    await state.set_state(UserStates.cart_change_qty_enter_new_qty)
    await callback.message.answer(f"Введите новое количество для товара **{display_name}** (текущее: {current_qty} шт):", parse_mode="Markdown")
    await callback.answer()

@router.message(UserStates.cart_change_qty_enter_new_qty)
async def cart_change_qty_process(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    try:
        new_quantity = int(message.text)
        if new_quantity < 0: raise ValueError()
        
        data = await state.get_data()
        characteristic_id = data.get('characteristic_id_to_change_qty') # ИЗМЕНЕНО: Получаем characteristic_id

        async with aiosqlite.connect("products.db") as db:
            cursor = await db.execute(
                """SELECT pc.name, pc.value, pc.quantity, p.name as product_name 
                   FROM product_characteristics pc 
                   JOIN products p ON pc.product_id = p.id
                   WHERE pc.id = ?""", (characteristic_id,)
            )
            result = await cursor.fetchone()
            if not result:
                await message.answer("Извините, этот вариант товара больше недоступен.")
                await state.clear()
                return

            char_name, char_value, available_qty, product_name = result
            display_name = f"{product_name} ({char_name}: {char_value})"

        if new_quantity > available_qty:
            await message.answer(f"Извините, доступно только {available_qty} шт. Пожалуйста, введите корректное количество.")
            return

        cart = await load_user_cart_from_db(user_id)
        
        if new_quantity == 0:
            cart = [(cid, qty) for cid, qty in cart if cid != characteristic_id]
        else:
            cart = [(cid, new_quantity if cid == characteristic_id else qty) for cid, qty in cart]

        await save_user_cart_to_db(user_id, cart)
        await state.clear()
        await message.answer(f"Количество товара '{display_name}' обновлено.")
        await view_cart(message)
    except ValueError:
        await message.answer("Пожалуйста, введите корректное число.")

@router.message(F.text == "✅ Оформить заказ")
async def start_order_process(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    cart = await load_user_cart_from_db(user_id)
    if not cart:
        await message.answer("Корзина пуста. Добавьте товары, прежде чем оформлять заказ.")
        return
    await state.set_state(UserStates.awaiting_delivery_name)
    await message.answer("Введите ваши фамилию и имя:")

@router.message(UserStates.awaiting_delivery_name)
async def process_delivery_name(message: types.Message, state: FSMContext):
    await state.update_data(delivery_name=message.text)
    await state.set_state(UserStates.awaiting_delivery_address)
    await message.answer("Введите ваш город, улицу, дом, квартиру (например: Москва, ул. Ленина, д. 10, кв. 5):")

@router.message(UserStates.awaiting_delivery_address)
async def process_delivery_address(message: types.Message, state: FSMContext):
    await state.update_data(delivery_address=message.text)
    await state.set_state(UserStates.awaiting_delivery_phone)
    await message.answer("Введите ваш номер телефона (например: +79123456789):")

@router.message(UserStates.awaiting_delivery_phone)
@router.message(UserStates.awaiting_delivery_phone)
async def process_delivery_phone(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    await state.update_data(delivery_phone=message.text)
    cart = await load_user_cart_from_db(user_id)
    data = await state.get_data()
    summary_lines = [f"**Данные доставки**:", f"Имя: {data['delivery_name']}", f"Адрес: {data['delivery_address']}", f"Телефон: {data['delivery_phone']}", "\n**Ваш заказ**:"]
    total_price = 0
    async with aiosqlite.connect("products.db") as db:
        for char_id, qty in cart: # Теперь получаем characteristic_id
            cursor = await db.execute(
                """SELECT pc.name, pc.value, pc.price, pc.quantity, p.name as product_name
                   FROM product_characteristics pc
                   JOIN products p ON pc.product_id = p.id
                   WHERE pc.id = ?""", (char_id,)
            )
            res = await cursor.fetchone()
            if not res or qty > res[3]: # res[3] - это quantity из product_characteristics
                display_name = f"{res[4] if res else 'Неизвестный товар'} ({res[0]}: {res[1]})" if res else f"ID {char_id}"
                await message.answer(f"Извините, товар '{display_name}' нет в достаточном количестве. Пожалуйста, измените корзину.")
                await state.clear()
                return
            
            char_name, char_value, char_price, available_qty, product_name = res
            item_total = char_price * qty / 100
            display_name_full = f"{product_name} ({char_name}: {char_value})"
            summary_lines.append(f"- {display_name_full} ({qty} шт) — {item_total:.2f}₽")
            total_price += item_total
    summary_lines.append(f"\n**Итого**: {total_price:.2f}₽")
    buttons = [[InlineKeyboardButton(text="✅ Подтвердить", callback_data="confirm_order"), InlineKeyboardButton(text="❌ Отменить", callback_data="cancel_order")]]
    await state.set_state(UserStates.awaiting_order_confirmation)
    await message.answer("\n".join(summary_lines), reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode="Markdown")


@router.callback_query(F.data == "cancel_order", UserStates.awaiting_order_confirmation)
async def cancel_order(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("Оформление заказа отменено.", reply_markup=None)
    await start(callback.message)
    await callback.answer()

# --- Админские функции ---

# <-- НАЧАЛО: НОВЫЕ ФУНКЦИИ ИМПОРТА/ЭКСПОРТА ТОВАРОВ -->

@router.message(F.text == "📤 Экспорт товаров", F.from_user.id == YOUR_ADMIN_ID)
async def export_products(message: types.Message):
    try:
        if message.from_user.id != YOUR_ADMIN_ID:
            await message.answer("У вас нет доступа к этой функции.")
            return

        async with aiosqlite.connect("products.db") as db:
            # Получаем все товары с их характеристиками
            cursor = await db.execute("""
                SELECT 
                    p.id as product_id,
                    p.name as product_name,
                    p.description,
                    pc.name as char_name,
                    pc.value as char_value,
                    pc.price,
                    pc.quantity
                FROM products p
                LEFT JOIN product_characteristics pc ON p.id = pc.product_id
                ORDER BY p.id, pc.id
            """)
            products = await cursor.fetchall()

            if not products:
                await message.answer("В базе данных нет товаров.")
                return

            # Формируем текст отчета
            current_time = datetime.utcnow().strftime("%d-%m-%Y")
            report_text = f"Отчет по товарам (UTC: {current_time})\n"
            report_text += f"Сформирован администратором: {message.from_user.full_name}\n\n"
            
            current_product_id = None
            total_items = 0
            total_value = 0

            for product in products:
                product_id, product_name, description, char_name, char_value, price, quantity = product
                
                if current_product_id != product_id:
                    current_product_id = product_id
                    report_text += f"Товар: {product_name}\n"
                    if description:
                        report_text += f"Описание: {description}\n"
                
                if char_name and char_value is not None:
                    price_rub = price / 100 if price else 0
                    report_text += f"- {char_name}: {char_value}\n"
                    report_text += f"  Цена: {price_rub:.2f} руб.\n"
                    report_text += f"  В наличии: {quantity} шт.\n"
                    
                    total_items += quantity if quantity else 0
                    total_value += (price_rub * quantity) if price and quantity else 0
                
                report_text += "\n"

            report_text += f"\nИтоговая статистика:\n"
            report_text += f"Всего единиц товара: {total_items} шт.\n"
            report_text += f"Общая стоимость: {total_value:.2f} руб."

            # Создаем временный файл отчета
            current_time_filename = datetime.utcnow().strftime("%d%m%Y")
            filename = f"products_export_{current_time_filename}.txt"
            
            with open(filename, "w", encoding="utf-8") as file:
                file.write(report_text)

            # Отправляем файл
            try:
                with open(filename, "rb") as file:
                    await message.answer_document(
                        document=types.BufferedInputFile(
                            file.read(),
                            filename=filename
                        ),
                        caption="Отчет по товарам сформирован"
                    )
            finally:
                # Удаляем временный файл
                if os.path.exists(filename):
                    os.remove(filename)

            # Отправляем краткую сводку
            summary = (
                f"Экспорт успешно завершен\n"
                f"Дата формирования (UTC): {current_time}\n"
                f"Всего позиций: {len(products)}\n"
                f"Общее количество: {total_items} шт.\n"
                f"Общая стоимость: {total_value:.2f} руб."
            )
            
            await message.answer(summary)

    except Exception as e:
        print(f"Error in export_products: {e}")
        await message.answer("Произошла ошибка при экспорте товаров. Пожалуйста, попробуйте позже.")


async def check_telegram_connection():
    """Проверяет соединение с серверами Telegram"""
    try:
        # Пробуем выполнить простой запрос к API Telegram
        await bot.get_me()
        return True
    except Exception as e:
        print(f"DEBUG: Telegram connection error: {e}")
        return False

# <-- КОНЕЦ: НОВЫЕ ФУНКЦИИ ИМПОРТА/ЭКСПОРТА ТОВАРОВ -->


@router.message(F.text == "📥 Заказы (админ)", F.from_user.id == YOUR_ADMIN_ID)
async def admin_orders(message: types.Message):
    try:
        async with aiosqlite.connect("products.db") as db:
            # Исправленный запрос для получения заказов
            cursor = await db.execute("""
                SELECT DISTINCT
                    o.id,
                    o.user_id,
                    o.status,
                    o.delivery_info,
                    (SELECT COUNT(*) FROM order_items WHERE order_id = o.id) as items_count,
                    (SELECT SUM(pc.price * oi2.quantity) 
                     FROM order_items oi2 
                     JOIN product_characteristics pc ON oi2.characteristic_id = pc.id 
                     WHERE oi2.order_id = o.id) as total_price
                FROM orders o
                ORDER BY o.id DESC
            """)
            orders = await cursor.fetchall()

            if not orders:
                await message.answer("Пока нет заказов.")
                return

            buttons = []
            for order in orders:
                order_id, user_id, status, delivery_info, items_count, total_price = order
                total_price = total_price / 100 if total_price else 0
                
                button_text = (f"Заказ #{order_id} | {status}\n"
                             f"От: ID {user_id} | {items_count} поз. | {total_price:.2f}₽")
                
                buttons.append([InlineKeyboardButton(
                    text=button_text,
                    callback_data=f"view_order_{order_id}"
                )])

            buttons.append([InlineKeyboardButton(
                text="↩️ Вернуться в админ-панель",
                callback_data="back_to_admin_panel"
            )])

            markup = InlineKeyboardMarkup(inline_keyboard=buttons)
            await message.answer("Список заказов:", reply_markup=markup)

    except Exception as e:
        print(f"Error in admin_orders: {e}")
        await message.answer("Произошла ошибка при получении списка заказов. Пожалуйста, попробуйте позже.")

@router.callback_query(F.data.startswith("status_"), F.from_user.id == YOUR_ADMIN_ID)
async def status_update(callback: types.CallbackQuery):
    order_id = int(callback.data.split('_')[1])
    statuses = ["Принят", "Отправлен", "Завершен"]
    buttons = [[InlineKeyboardButton(text=s, callback_data=f"set_{order_id}_{s}")] for s in statuses]
    await callback.message.answer("Выберите новый статус:", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

@router.callback_query(F.data.startswith("set_"), F.from_user.id == YOUR_ADMIN_ID)
async def set_status(callback: types.CallbackQuery):
    _, order_id, new_status = callback.data.split("_", 2)
    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE orders SET status = ? WHERE id = ?", (new_status, int(order_id)))
        cursor = await db.execute("SELECT user_id FROM orders WHERE id = ?", (int(order_id),))
        user = await cursor.fetchone()
        await db.commit()
    if user:
        try:
            await bot.send_message(user[0], f"Статус вашего заказа #{order_id} изменён на: {new_status}")
        except Exception as e:
            await notify_admin(f"Не удалось уведомить пользователя {user[0]} о смене статуса заказа #{order_id}: {e}")
    await callback.answer(f"Статус заказа #{order_id} обновлён на {new_status}")
    await admin_orders(callback.message)



# Открытие меню добавления товара
@router.message(F.text == "➕ Добавить товар", F.from_user.id == YOUR_ADMIN_ID)
async def cmd_add_product(message: types.Message, state: FSMContext):
    await state.set_state(AdminStates.add_product_name)
    await message.answer("Введите **название** нового товара:", parse_mode="Markdown")

# Обработка названия товара
@router.message(AdminStates.add_product_name)
async def process_add_product_name(message: types.Message, state: FSMContext):
    await state.update_data(product_name=message.text)
    await state.set_state(AdminStates.add_product_description)
    await message.answer("Введите **описание** товара:")

# Обработка описания товара
@router.message(AdminStates.add_product_description)
async def process_add_product_description(message: types.Message, state: FSMContext):
    await state.update_data(product_description=message.text)
    await state.set_state(AdminStates.add_product_image)
    await message.answer("Теперь отправьте **фотографию** товара или введите 'без фото', если ее нет:", parse_mode="Markdown")

# Обработка изображения товара
@router.message(AdminStates.add_product_image)
async def process_add_product_image(message: types.Message, state: FSMContext):
    image_url = None
    if message.photo:
        image_url = message.photo[-1].file_id # Берем самое большое фото
    elif message.text and message.text.lower() == 'без фото':
        image_url = None
    else:
        await message.answer("Пожалуйста, отправьте фотографию или введите 'без фото'.")
        return

    await state.update_data(product_image_url=image_url)

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT id, name FROM categories")
        categories = await cursor.fetchall()

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=name, callback_data=f"select_cat_{cat_id}")] for cat_id, name in categories
    ] + [[InlineKeyboardButton(text="Отмена", callback_data="cancel_add_product_category")]])
    
    await state.set_state(AdminStates.add_product_category)
    await message.answer("Выберите **категорию** товара:", reply_markup=markup, parse_mode="Markdown")

# Обработка выбора категории и добавление базового товара
@router.callback_query(AdminStates.add_product_category, F.data.startswith("select_cat_"))
async def process_add_product_category_selected(callback: types.CallbackQuery, state: FSMContext):
    category_id = int(callback.data.split("_")[2])
    data = await state.get_data()
    name = data['product_name']
    description = data['product_description']
    image_url = data['product_image_url']

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute(
            "INSERT INTO products (name, description, category_id, image_url) VALUES (?, ?, ?, ?)",
            (name, description, category_id, image_url)
        )
        product_id = cursor.lastrowid
        await db.commit()

    await state.update_data(current_product_id=product_id) # Сохраняем ID нового товара
    await callback.message.edit_text(f"Товар '{name}' успешно добавлен. Теперь добавьте характеристики:", reply_markup=None)
    
    # Переходим к добавлению характеристик
    await state.set_state(AdminStates.add_characteristic_name)
    await callback.message.answer(
        "Введите **название характеристики** (например, 'Объем', 'Цвет').\n"
        "Для завершения добавления характеристик введите 'готово'.", parse_mode="Markdown"
    )
    await callback.answer()

@router.callback_query(AdminStates.add_product_category, F.data == "cancel_add_product_category")
async def cancel_add_product_category(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("Добавление товара отменено.", reply_markup=None)
    await callback.answer()
    await admin_account(callback.message) # Возвращаем в админ-панель


# --- Добавление характеристик к товару ---

@router.message(AdminStates.add_characteristic_name)
async def process_add_characteristic_name(message: types.Message, state: FSMContext):
    if message.text.lower() == 'готово':
        await state.clear()
        await message.answer("Добавление характеристик завершено.")
        await admin_account(message)
        return

    await state.update_data(current_characteristic_name=message.text)
    await state.set_state(AdminStates.add_characteristic_value)
    await message.answer(f"Введите **значение характеристики** '{message.text}' (например, '5мл', 'Красный'):", parse_mode="Markdown")

@router.message(AdminStates.add_characteristic_value)
async def process_add_characteristic_value(message: types.Message, state: FSMContext):
    await state.update_data(current_characteristic_value=message.text)
    await state.set_state(AdminStates.add_characteristic_price)
    await message.answer("Введите **цену** для этой характеристики (в рублях, например '1200.50'):", parse_mode="Markdown")

@router.message(AdminStates.add_characteristic_price)
async def process_add_characteristic_price(message: types.Message, state: FSMContext):
    try:
        price_rub = float(message.text.replace(',', '.'))
        price_kopecks = int(price_rub * 100)
        if price_kopecks <= 0: raise ValueError
        await state.update_data(current_characteristic_price=price_kopecks)
        await state.set_state(AdminStates.add_characteristic_quantity)
        await message.answer("Введите **количество** для этой характеристики (целое число):", parse_mode="Markdown")
    except ValueError:
        await message.answer("Пожалуйста, введите корректную цену (например, 1200.50).")

@router.message(AdminStates.add_characteristic_quantity)
async def process_add_characteristic_quantity(message: types.Message, state: FSMContext):
    try:
        quantity = int(message.text)
        if quantity < 0: raise ValueError
        
        data = await state.get_data()
        product_id = data['current_product_id']
        char_name = data['current_characteristic_name']
        char_value = data['current_characteristic_value']
        char_price = data['current_characteristic_price']

        async with aiosqlite.connect("products.db") as db:
            await db.execute(
                "INSERT INTO product_characteristics (product_id, name, value, price, quantity) VALUES (?, ?, ?, ?, ?)",
                (product_id, char_name, char_value, char_price, quantity)
            )
            await db.commit()

        await message.answer(f"Характеристика '{char_name}: {char_value}' (Цена: {char_price / 100:.2f}₽, Кол-во: {quantity}) успешно добавлена.")
        
        # Предлагаем добавить еще характеристику или завершить
        await state.set_state(AdminStates.add_characteristic_name) # Возвращаемся к вводу названия новой характеристики
        await message.answer(
            "Введите **название следующей характеристики** или 'готово' для завершения:", parse_mode="Markdown"
        )

    except ValueError:
        await message.answer("Пожалуйста, введите корректное количество (целое неотрицательное число).")


# --- Управление характеристиками существующих товаров ---

@router.message(F.text == "🎛️ Упр. характеристиками", F.from_user.id == YOUR_ADMIN_ID)
async def manage_characteristics_start(message: types.Message, state: FSMContext):
    await state.set_state(AdminStates.manage_product_characteristics)
    await message.answer("Выберите товар, характеристики которого хотите изменить:", 
                         reply_markup=await get_product_selection_markup("manage_char"))

# Вспомогательная функция для генерации кнопок выбора товара
async def get_product_selection_markup(callback_prefix: str) -> InlineKeyboardMarkup:
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT id, name FROM products ORDER BY name")
        products = await cursor.fetchall()
    
    buttons = []
    for prod_id, prod_name in products:
        buttons.append([InlineKeyboardButton(text=prod_name, callback_data=f"{callback_prefix}_{prod_id}")])
    buttons.append([InlineKeyboardButton(text="Отмена", callback_data=f"{callback_prefix}_cancel")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

@router.callback_query(AdminStates.manage_product_characteristics, F.data.startswith("manage_char_"))
async def select_product_for_characteristic_management(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    action = parts[2] # 'prod' or 'cancel'

    if action == "cancel":
        await state.clear()
        await callback.message.edit_text("Управление характеристиками отменено.", reply_markup=None)
        await callback.answer()
        await admin_account(callback.message)
        return

    product_id = int(parts[2]) # ID товара
    await state.update_data(current_product_id=product_id)

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name FROM products WHERE id = ?", (product_id,))
        product_name = (await cursor.fetchone())[0]
        
        cursor = await db.execute("SELECT id, name, value, price, quantity FROM product_characteristics WHERE product_id = ?", (product_id,))
        characteristics = await cursor.fetchall()
    
    await callback.message.edit_text(f"Управление характеристиками для товара: **{product_name}**", parse_mode="Markdown", reply_markup=None)

    char_buttons = []
    if characteristics:
        for char_id, char_name, char_value, char_price, char_qty in characteristics:
            char_buttons.append([
                InlineKeyboardButton(text=f"✏️ {char_name}: {char_value} ({char_price/100:.2f}₽, {char_qty}шт)", callback_data=f"edit_char_{char_id}"),
                InlineKeyboardButton(text=f"🗑️", callback_data=f"del_char_{char_id}")
            ])
    
    char_buttons.append([InlineKeyboardButton(text="➕ Добавить новую характеристику", callback_data=f"add_new_char_{product_id}")])
    char_buttons.append([InlineKeyboardButton(text="⬅️ Назад к выбору товара", callback_data="manage_char_back_to_products")])
    char_buttons.append([InlineKeyboardButton(text="↩️ В админ-панель", callback_data="back_to_admin_panel")])

    await callback.message.answer("Выберите характеристику для редактирования или добавьте новую:", 
                                  reply_markup=InlineKeyboardMarkup(inline_keyboard=char_buttons))
    await callback.answer()

@router.callback_query(F.data == "manage_char_back_to_products", AdminStates.manage_product_characteristics)
async def back_to_product_selection_for_char_manage(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(AdminStates.manage_product_characteristics) # Снова устанавливаем состояние
    await callback.message.edit_text("Выберите товар, характеристики которого хотите изменить:", 
                                     reply_markup=await get_product_selection_markup("manage_char"))
    await callback.answer()

@router.callback_query(F.data == "back_to_admin_panel", AdminStates.manage_product_characteristics)
async def back_to_admin_panel_from_char_manage(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete()
    await admin_account(callback.message)
    await callback.answer()

@router.callback_query(F.data.startswith("add_new_char_"))
async def add_new_characteristic_callback(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[3])
    await state.update_data(current_product_id=product_id) # Убедимся, что product_id в данных состояния
    await state.set_state(AdminStates.add_characteristic_name) # Переходим к общему flow добавления характеристики
    await callback.message.edit_text(
        "Введите **название характеристики** (например, 'Объем', 'Цвет').\n"
        "Для завершения добавления характеристик введите 'готово'.", parse_mode="Markdown", reply_markup=None
    )
    await callback.answer()

@router.callback_query(F.data.startswith("edit_char_"))
async def edit_characteristic_select_field(callback: types.CallbackQuery, state: FSMContext):
    characteristic_id = int(callback.data.split("_")[2])
    await state.update_data(current_characteristic_id=characteristic_id)

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name, value, price, quantity FROM product_characteristics WHERE id = ?", (characteristic_id,))
        char_name, char_value, char_price, char_qty = await cursor.fetchone()
    
    await state.update_data(
        original_char_name=char_name, 
        original_char_value=char_value,
        original_char_price=char_price,
        original_char_quantity=char_qty
    )

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"Изменить название ({char_name})", callback_data="edit_char_field_name")],
        [InlineKeyboardButton(text=f"Изменить значение ({char_value})", callback_data="edit_char_field_value")],
        [InlineKeyboardButton(text=f"Изменить цену ({char_price/100:.2f}₽)", callback_data="edit_char_field_price")],
        [InlineKeyboardButton(text=f"Изменить количество ({char_qty}шт)", callback_data="edit_char_field_quantity")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data=f"back_to_char_list_{data['current_product_id']}")] # Возврат к списку характеристик
    ])
    await callback.message.edit_text("Что вы хотите изменить в этой характеристике?", reply_markup=markup)
    await callback.answer()

@router.callback_query(F.data.startswith("back_to_char_list_"))
async def back_to_characteristic_list(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[3])
    await state.update_data(current_product_id=product_id) # Убедимся, что product_id в данных состояния
    # Используем ту же логику, что и при первом выборе товара для управления характеристиками
    await select_product_for_characteristic_management(callback, state) # Переиспользуем функцию
    await callback.answer()


@router.callback_query(F.data.startswith("edit_char_field_"))
async def edit_characteristic_field_start(callback: types.CallbackQuery, state: FSMContext):
    field_to_edit = callback.data.split("_")[3] # 'name', 'value', 'price', 'quantity'
    await state.update_data(field_to_edit=field_to_edit)

    data = await state.get_data()
    original_value = data[f'original_char_{field_to_edit}']

    if field_to_edit == "price":
        await state.set_state(AdminStates.edit_characteristic_price_input)
        await callback.message.edit_text(f"Введите новую **цену** ({original_value/100:.2f}₽):", parse_mode="Markdown")
    elif field_to_edit == "quantity":
        await state.set_state(AdminStates.edit_characteristic_quantity_input)
        await callback.message.edit_text(f"Введите новое **количество** ({original_value} шт):", parse_mode="Markdown")
    else:
        await state.set_state(AdminStates.edit_characteristic_field)
        await callback.message.edit_text(f"Введите новое **{field_to_edit}** (текущее: {original_value}):", parse_mode="Markdown")
    
    await callback.answer()

@router.message(AdminStates.edit_characteristic_field)
async def process_edit_characteristic_field(message: types.Message, state: FSMContext):
    data = await state.get_data()
    characteristic_id = data['current_characteristic_id']
    field_to_edit = data['field_to_edit']
    new_value = message.text

    async with aiosqlite.connect("products.db") as db:
        query = f"UPDATE product_characteristics SET {field_to_edit} = ? WHERE id = ?"
        await db.execute(query, (new_value, characteristic_id))
        await db.commit()
    
    await message.answer(f"'{field_to_edit}' успешно обновлено на: {new_value}.")
    await state.clear()
    # Возвращаемся к списку характеристик этого товара
    product_id = data['current_product_id']
    callback_for_return = types.CallbackQuery(id='dummy', from_user=message.from_user, chat_instance='dummy', data=f"dummy_callback_data")
    callback_for_return.message = message # Привязываем текущее сообщение к callback
    await back_to_characteristic_list(callback_for_return, state)


@router.message(AdminStates.edit_characteristic_price_input)
async def process_edit_characteristic_price_input(message: types.Message, state: FSMContext):
    try:
        new_price_rub = float(message.text.replace(',', '.'))
        new_price_kopecks = int(new_price_rub * 100)
        if new_price_kopecks <= 0: raise ValueError

        data = await state.get_data()
        characteristic_id = data['current_characteristic_id']

        async with aiosqlite.connect("products.db") as db:
            await db.execute("UPDATE product_characteristics SET price = ? WHERE id = ?", (new_price_kopecks, characteristic_id))
            await db.commit()
        
        await message.answer(f"Цена успешно обновлена на: {new_price_rub:.2f}₽.")
        await state.clear()
        product_id = data['current_product_id']
        callback_for_return = types.CallbackQuery(id='dummy', from_user=message.from_user, chat_instance='dummy', data=f"dummy_callback_data")
        callback_for_return.message = message
        await back_to_characteristic_list(callback_for_return, state)
    except ValueError:
        await message.answer("Пожалуйста, введите корректную цену (например, 1200.50).")

@router.message(AdminStates.edit_characteristic_quantity_input)
async def process_edit_characteristic_quantity_input(message: types.Message, state: FSMContext):
    try:
        new_quantity = int(message.text)
        if new_quantity < 0: raise ValueError

        data = await state.get_data()
        characteristic_id = data['current_characteristic_id']

        async with aiosqlite.connect("products.db") as db:
            await db.execute("UPDATE product_characteristics SET quantity = ? WHERE id = ?", (new_quantity, characteristic_id))
            await db.commit()
        
        await message.answer(f"Количество успешно обновлено на: {new_quantity} шт.")
        await state.clear()
        product_id = data['current_product_id']
        callback_for_return = types.CallbackQuery(id='dummy', from_user=message.from_user, chat_instance='dummy', data=f"dummy_callback_data")
        callback_for_return.message = message
        await back_to_characteristic_list(callback_for_return, state)
    except ValueError:
        await message.answer("Пожалуйста, введите корректное количество (целое неотрицательное число).")

@router.callback_query(F.data.startswith("del_char_"))
async def delete_characteristic(callback: types.CallbackQuery, state: FSMContext):
    characteristic_id = int(callback.data.split("_")[2])
    data = await state.get_data()
    product_id = data['current_product_id']

    async with aiosqlite.connect("products.db") as db:
        await db.execute("DELETE FROM product_characteristics WHERE id = ?", (characteristic_id,))
        await db.commit()
    
    await callback.message.edit_text("Характеристика успешно удалена.", reply_markup=None)
    
    # Возвращаемся к списку характеристик
    await state.update_data(current_product_id=product_id) # Убедимся, что product_id в данных состояния
    await select_product_for_characteristic_management(callback, state)
    await callback.answer()

# --- Удаление товара (требует изменений, так как удаление товара должно удалять и его характеристики) ---
@router.callback_query(F.data.startswith("delete_prod_")) # Обновляем этот обработчик
async def delete_product_start(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split("_")[2])
    await state.update_data(product_id_to_delete=product_id)
    
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name FROM products WHERE id = ?", (product_id,))
        product_name = (await cursor.fetchone())[0]

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да, удалить", callback_data="confirm_delete_product")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_delete_product")]
    ])
    await state.set_state(AdminStates.delete_product_confirmation) # Устанавливаем состояние подтверждения
    await callback.message.answer(f"Вы уверены, что хотите удалить товар **'{product_name}'** и все его характеристики?", reply_markup=markup, parse_mode="Markdown")
    await callback.answer()

@router.callback_query(F.data == "confirm_delete_product", AdminStates.delete_product_confirmation)
async def confirm_delete_product(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    product_id_to_delete = data.get('product_id_to_delete')

    async with aiosqlite.connect("products.db") as db:
        # Благодаря FOREIGN KEY ON DELETE CASCADE в product_characteristics,
        # удаление товара автоматически удалит все его характеристики.
        await db.execute("DELETE FROM products WHERE id = ?", (product_id_to_delete,))
        await db.commit()

    await state.clear()
    await callback.message.edit_text("Товар и все его характеристики успешно удалены.", reply_markup=None)
    # Возвращаемся в меню редактирования товаров или админ-панель
    # Для простоты, вернемся в админ-панель
    await admin_account(callback.message)
    await callback.answer()

@router.callback_query(F.data == "cancel_delete_product", AdminStates.delete_product_confirmation)
async def cancel_delete_product(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("Удаление товара отменено.", reply_markup=None)
    # Возвращаемся в меню редактирования товаров или админ-панель
    await admin_account(callback.message)
    await callback.answer()

@router.message(F.text == "✏️ Редактировать товар", F.from_user.id == YOUR_ADMIN_ID)
async def edit_product_select(message: types.Message, state: FSMContext):
    await state.set_state(AdminStates.edit_product_select)
    await message.answer("Выберите товар для редактирования:", reply_markup=await get_product_selection_markup("edit_prod"))

@router.callback_query(AdminStates.edit_product_select, F.data.startswith("edit_prod_"))
async def process_edit_product_select(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    action = parts[2]

    if action == "cancel":
        await state.clear()
        await callback.message.edit_text("Редактирование товара отменено.", reply_markup=None)
        await callback.answer()
        await admin_account(callback.message)
        return

    product_id = int(parts[2])
    await state.update_data(product_id_to_edit=product_id)

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name, description, category_id, image_url FROM products WHERE id = ?", (product_id,))
        name, description, category_id, image_url = await cursor.fetchone()
        category_name = await get_category_name(category_id)
        
        # Получаем информацию о характеристиках для отображения
        cursor_chars = await db.execute("SELECT name, value, price, quantity FROM product_characteristics WHERE product_id = ?", (product_id,))
        characteristics = await cursor_chars.fetchall()

    char_info = ""
    if characteristics:
        char_info = "\n\n**Характеристики:**\n"
        for cn, cv, cp, cq in characteristics:
            char_info += f"- {cn}: {cv} ({cp/100:.2f}₽, {cq} шт)\n"
    else:
        char_info = "\n\n**Нет характеристик. Товар недоступен для заказа.**"

    text = (f"**Текущие данные товара:**\n"
            f"Название: {name}\n"
            f"Описание: {description or 'Нет описания'}\n"
            f"Категория: {category_name}\n"
            f"URL Изображения: {image_url or 'Нет'}{char_info}")

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Изменить название", callback_data="edit_field_name")],
        [InlineKeyboardButton(text="Изменить описание", callback_data="edit_field_description")],
        [InlineKeyboardButton(text="Изменить категорию", callback_data="edit_field_category")],
        [InlineKeyboardButton(text="Изменить изображение", callback_data="edit_field_image")],
        [InlineKeyboardButton(text="🗑️ Удалить товар", callback_data=f"delete_prod_{product_id}")], # Передача product_id
        [InlineKeyboardButton(text="↩️ В админ-панель", callback_data="back_to_admin_panel_from_edit_product")]
    ])
    await state.set_state(AdminStates.edit_product_menu)
    await callback.message.edit_text(text, reply_markup=markup, parse_mode="Markdown")
    await callback.answer()

@router.callback_query(F.data == "back_to_admin_panel_from_edit_product", AdminStates.edit_product_menu)
async def back_to_admin_panel_from_edit_product(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete()
    await admin_account(callback.message)
    await callback.answer()


@router.callback_query(AdminStates.edit_product_menu, F.data.startswith("edit_field_"))
async def edit_product_field(callback: types.CallbackQuery, state: FSMContext):
    field = callback.data.split("_")[2]
    await state.update_data(edit_field=field)

    data = await state.get_data()
    product_id = data['product_id_to_edit']

    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name, description, category_id, image_url FROM products WHERE id = ?", (product_id,))
        product_data = await cursor.fetchone()
        
        current_value = ""
        if field == "name":
            current_value = product_data[0]
            await state.set_state(AdminStates.edit_product_name_input)
            await callback.message.edit_text(f"Введите новое название товара (текущее: {current_value}):")
        elif field == "description":
            current_value = product_data[1]
            await state.set_state(AdminStates.edit_product_description_input)
            await callback.message.edit_text(f"Введите новое описание товара (текущее: {current_value or 'нет описания'}):")
        elif field == "category":
            cursor_cat = await db.execute("SELECT id, name FROM categories")
            categories = await cursor_cat.fetchall()
            markup = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=name, callback_data=f"select_edit_cat_{cat_id}")] for cat_id, name in categories
            ] + [[InlineKeyboardButton(text="Отмена", callback_data="cancel_edit_product_category")]])
            await state.set_state(AdminStates.edit_product_category_input)
            await callback.message.edit_text("Выберите новую категорию:", reply_markup=markup)
        elif field == "image":
            current_value = product_data[3]
            await state.set_state(AdminStates.edit_product_image_input)
            await callback.message.edit_text(f"Отправьте новое изображение товара или введите 'без фото' (текущее URL: {current_value or 'нет'}):")
    await callback.answer()

@router.message(AdminStates.edit_product_name_input)
async def process_edit_product_name_input(message: types.Message, state: FSMContext):
    data = await state.get_data()
    product_id = data['product_id_to_edit']
    new_name = message.text
    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE products SET name = ? WHERE id = ?", (new_name, product_id))
        await db.commit()
    await message.answer("Название товара обновлено.")
    await state.clear()
    await edit_product_select(message, state) # Возвращаемся к выбору товара для редактирования


@router.message(AdminStates.edit_product_description_input)
async def process_edit_product_description_input(message: types.Message, state: FSMContext):
    data = await state.get_data()
    product_id = data['product_id_to_edit']
    new_description = message.text
    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE products SET description = ? WHERE id = ?", (new_description, product_id))
        await db.commit()
    await message.answer("Описание товара обновлено.")
    await state.clear()
    await edit_product_select(message, state)


@router.callback_query(AdminStates.edit_product_category_input, F.data.startswith("select_edit_cat_"))
async def process_edit_product_category_selected(callback: types.CallbackQuery, state: FSMContext):
    category_id = int(callback.data.split("_")[3])
    data = await state.get_data()
    product_id = data['product_id_to_edit']
    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE products SET category_id = ? WHERE id = ?", (category_id, product_id))
        await db.commit()
    await callback.message.edit_text("Категория товара обновлена.", reply_markup=None)
    await state.clear()
    await edit_product_select(callback.message, state)
    await callback.answer()

@router.callback_query(AdminStates.edit_product_category_input, F.data == "cancel_edit_product_category")
async def cancel_edit_product_category(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("Изменение категории отменено.", reply_markup=None)
    await callback.answer()
    await edit_product_select(callback.message, state) # Возвращаемся к выбору товара для редактирования

@router.message(AdminStates.edit_product_image_input)
async def process_edit_product_image_input(message: types.Message, state: FSMContext):
    data = await state.get_data()
    product_id = data['product_id_to_edit']
    image_url = None
    if message.photo:
        image_url = message.photo[-1].file_id
    elif message.text and message.text.lower() == 'без фото':
        image_url = None
    else:
        await message.answer("Пожалуйста, отправьте фотографию или введите 'без фото'.")
        return
    
    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE products SET image_url = ? WHERE id = ?", (image_url, product_id))
        await db.commit()
    await message.answer("Изображение товара обновлено.")
    await state.clear()
    await edit_product_select(message, state)

@router.callback_query(F.data == "delete_product_image", AdminStates.edit_product_image_input, F.from_user.id == YOUR_ADMIN_ID)
async def edit_product_image_process_delete(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    product_id = data.get('product_id_to_edit')
    if not product_id:
        await callback.message.answer("Произошла ошибка. Пожалуйста, попробуйте снова.")
        await state.clear()
        return
    
    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE products SET image_url = NULL WHERE id = ?", (product_id,))
        await db.commit()
    await callback.message.edit_text("Изображение товара успешно удалено!")
    await state.clear()
    await callback.answer()

@router.callback_query(F.data == "skip_image_edit", AdminStates.edit_product_image_input, F.from_user.id == YOUR_ADMIN_ID)
async def edit_product_image_process_skip(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Изменение изображения пропущено. Текущее изображение сохранено.")
    await state.clear()
    await callback.answer()

@router.message(AdminStates.edit_product_image_input, F.from_user.id == YOUR_ADMIN_ID) # Обработка любого другого текста в состоянии изменения изображения
async def edit_product_image_invalid_input(message: types.Message, state: FSMContext):
    markup_buttons = [[InlineKeyboardButton(text="Пропустить (оставить как есть)", callback_data="skip_image_edit")]]
    data = await state.get_data()
    current_image_url = data.get('current_image_url')
    if current_image_url:
        markup_buttons.insert(0, [InlineKeyboardButton(text="Удалить изображение", callback_data="delete_product_image")])

    markup = InlineKeyboardMarkup(inline_keyboard=markup_buttons)
    await message.answer("Пожалуйста, отправьте новое изображение, нажмите 'Удалить изображение' или 'Пропустить'.", reply_markup=markup)

@router.callback_query(F.data == "cancel_edit", AdminStates.edit_product_menu, F.from_user.id == YOUR_ADMIN_ID)
async def cancel_edit_product(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("Редактирование товара отменено.")
    await callback.answer()

@router.callback_query(F.data == "delete_product", AdminStates.edit_product_menu, F.from_user.id == YOUR_ADMIN_ID)
async def delete_product_confirm_prompt(callback: types.CallbackQuery, state: FSMContext):
    """Запрашивает подтверждение на удаление товара."""
    data = await state.get_data()
    product_name = data.get('current_name', 'Неизвестный товар')

    # Меняем состояние на ожидание подтверждения
    await state.set_state(AdminStates.delete_product_confirmation)

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да, удалить", callback_data="delete_product_confirm")],
        [InlineKeyboardButton(text="❌ Нет, отмена", callback_data="cancel_edit")] # Можно использовать тот же колбэк отмены
    ])

    await callback.message.edit_text(
        f"Вы уверены, что хотите удалить товар **{product_name}**?\n\n"
        f"⚠️ **Это действие необратимо!**",
        reply_markup=markup,
        parse_mode="Markdown"
    )
    await callback.answer()

@router.callback_query(F.data == "delete_product_confirm", AdminStates.delete_product_confirmation, F.from_user.id == YOUR_ADMIN_ID)
async def delete_product_execute(callback: types.CallbackQuery, state: FSMContext):
    """Выполняет удаление товара после подтверждения."""
    data = await state.get_data()
    product_id = data.get('product_id_to_edit')
    product_name = data.get('current_name', 'Неизвестный товар')

    if not product_id:
        await callback.message.edit_text("Произошла ошибка. Не удалось определить товар для удаления. Попробуйте снова.")
        await state.clear()
        await callback.answer(show_alert=True, text="Ошибка!")
        return

    async with aiosqlite.connect("products.db") as db:
        await db.execute("DELETE FROM products WHERE id = ?", (product_id,))
        await db.commit()
    
    await state.clear()
    await callback.message.edit_text(f"✅ Товар **{product_name}** был успешно удален.", parse_mode="Markdown")
    await callback.answer()    
    
# --- НОВЫЙ БЛОК: УПРАВЛЕНИЕ КАТЕГОРИЯМИ ---

async def get_categories_markup():
    """Создает клавиатуру со списком категорий для управления."""
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT id, name FROM categories ORDER BY name")
        categories = await cursor.fetchall()
    
    buttons = []
    if categories:
        for cat_id, cat_name in categories:
            buttons.append([InlineKeyboardButton(text=cat_name, callback_data=f"select_cat_manage_{cat_id}")])
    
    buttons.append([InlineKeyboardButton(text="➕ Добавить новую категорию", callback_data="add_new_category")])
    buttons.append([InlineKeyboardButton(text="⬅️ Назад в админ-панель", callback_data="back_to_admin_panel")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


@router.message(F.text == "🗂️ Категории товаров", F.from_user.id == YOUR_ADMIN_ID)
async def manage_categories_menu(message: types.Message, state: FSMContext):
    """Отображает главное меню управления категориями."""
    await state.clear()
    markup = await get_categories_markup()
    await message.answer("Здесь вы можете управлять категориями товаров:", reply_markup=markup)


@router.callback_query(F.data == "back_to_cat_list")
async def back_to_categories_list(callback: types.CallbackQuery, state: FSMContext):
    """Возвращает к списку категорий из меню действий."""
    await state.clear()
    markup = await get_categories_markup()
    await callback.message.edit_text("Выберите категорию для управления:", reply_markup=markup)
    await callback.answer()

@router.callback_query(F.data == "back_to_admin_panel")
async def back_to_admin_panel_from_cats(callback: types.CallbackQuery, state: FSMContext):
    """Возвращает в главное меню админа."""
    await callback.message.delete()
    await admin_account(callback.message) # Вызываем функцию, которая строит админ-меню
    await callback.answer()


# --- Добавление категории ---
@router.callback_query(F.data == "add_new_category")
async def add_category_prompt(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(AdminStates.add_category_name)
    await callback.message.edit_text("Введите название для новой категории:")
    await callback.answer()

@router.message(AdminStates.add_category_name)
async def add_category_process(message: types.Message, state: FSMContext):
    new_name = message.text.strip()
    async with aiosqlite.connect("products.db") as db:
        await db.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (new_name,))
        await db.commit()
    
    await message.answer(f"Категория '{new_name}' успешно добавлена.")
    markup = await get_categories_markup()
    await message.answer("Меню категорий обновлено:", reply_markup=markup)
    await state.clear()


# --- Выбор действия с категорией (Переименовать/Удалить) ---
@router.callback_query(F.data.startswith("select_cat_manage_"))
async def category_action_menu(callback: types.CallbackQuery, state: FSMContext):
    category_id = int(callback.data.split("_")[3])
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT name FROM categories WHERE id = ?", (category_id,))
        category_name = (await cursor.fetchone())[0]
    
    await state.update_data(category_id_to_manage=category_id, category_name_to_manage=category_name)
    
    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ Переименовать", callback_data="rename_category")],
        [InlineKeyboardButton(text="🗑️ Удалить", callback_data="delete_category")],
        [InlineKeyboardButton(text="⬅️ Назад к списку", callback_data="back_to_cat_list")]
    ])
    await callback.message.edit_text(f"Выберите действие для категории '{category_name}':", reply_markup=markup)
    await callback.answer()


# --- Переименование категории ---
@router.callback_query(F.data == "rename_category")
async def rename_category_prompt(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    category_name = data.get('category_name_to_manage')
    await state.set_state(AdminStates.rename_category_name)
    await callback.message.edit_text(f"Введите новое название для категории '{category_name}':")
    await callback.answer()

@router.message(AdminStates.rename_category_name)
async def rename_category_process(message: types.Message, state: FSMContext):
    new_name = message.text.strip()
    data = await state.get_data()
    category_id = data.get('category_id_to_manage')

    async with aiosqlite.connect("products.db") as db:
        await db.execute("UPDATE categories SET name = ? WHERE id = ?", (new_name, category_id))
        await db.commit()
    
    await message.answer(f"Категория успешно переименована в '{new_name}'.")
    markup = await get_categories_markup()
    await message.answer("Меню категорий обновлено:", reply_markup=markup)
    await state.clear()


# --- Удаление категории ---
@router.callback_query(F.data == "delete_category")
async def delete_category_confirm_prompt(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    category_id = data.get('category_id_to_manage')
    category_name = data.get('category_name_to_manage')

    # Узнаем, сколько товаров находится в этой категории
    async with aiosqlite.connect("products.db") as db:
        cursor = await db.execute("SELECT COUNT(*) FROM products WHERE category_id = ?", (category_id,))
        product_count = (await cursor.fetchone())[0]

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да, удалить", callback_data="delete_category_confirm")],
        [InlineKeyboardButton(text="❌ Нет, отмена", callback_data="back_to_cat_list")]
    ])
    
    await callback.message.edit_text(
        f"Вы уверены, что хотите удалить категорию '{category_name}'?\n\n"
        f"В этой категории находится **{product_count}** товаров. После удаления они станут 'Без категории'.\n\n"
        f"⚠️ **Это действие необратимо!**",
        reply_markup=markup,
        parse_mode="Markdown"
    )
    await callback.answer()


@router.callback_query(F.data == "delete_category_confirm")
async def delete_category_execute(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    category_id = data.get('category_id_to_manage')
    category_name = data.get('category_name_to_manage')

    async with aiosqlite.connect("products.db") as db:
        # 1. Все товары из этой категории делаем "Без категории"
        await db.execute("UPDATE products SET category_id = NULL WHERE category_id = ?", (category_id,))
        # 2. Удаляем саму категорию
        await db.execute("DELETE FROM categories WHERE id = ?", (category_id,))
        await db.commit()

    await state.clear()
    await callback.message.edit_text(f"Категория '{category_name}' успешно удалена.")
    
    markup = await get_categories_markup()
    await callback.message.answer("Меню категорий обновлено:", reply_markup=markup)
    await callback.answer()


  # # Регистрация роутера
# dp.include_router(router)  

async def handle(request):
    return web.Response(text="Бот работает!")

async def start_bot(app):
    try:
        await init_db()
        asyncio.create_task(clear_old_carts_task())
        asyncio.create_task(dp.start_polling(bot))
        print("✅ Бот запущен в фоновом режиме")
    except Exception as e:
        print(f"🚨 Ошибка запуска бота: {e}")

if __name__ == '__main__':
    app = web.Application()
    app.router.add_get('/', handle)
    app.on_startup.append(start_bot)
    port = int(os.environ.get("PORT", 10000))
    print(f"🚀 Запуск веб-сервера на порту {port}")
    web.run_app(app, host='0.0.0.0', port=port)
    
